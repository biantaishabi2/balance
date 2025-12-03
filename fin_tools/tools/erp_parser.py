# -*- coding: utf-8 -*-
"""
erp_parser.py - ERP 数据解析器

支持金蝶、用友、SAP 等主流财务软件导出的 Excel/CSV 文件解析，
转换为项目统一的标准 JSON 格式。

支持的模板：
- trial_balance: 科目余额表
- budget: 预算对比表
- cash: 现金流数据
- voucher: 记账凭证
- ar_aging: 应收账龄表
- ap_aging: 应付账龄表
"""

from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
from datetime import datetime
import re


# ============================================================
# 列名映射配置
# ============================================================

# 通用列名映射（支持金蝶/用友/SAP 等多种 ERP）
COLUMN_MAPPING = {
    # 科目编码
    "科目代码": "account_code",
    "科目编码": "account_code",
    "会计科目": "account_code",
    "科目": "account_code",
    "GL Account": "account_code",
    "Account Code": "account_code",
    "账户编号": "account_code",

    # 科目名称
    "科目名称": "account_name",
    "科目全称": "account_name",
    "名称": "account_name",
    "Description": "account_name",
    "Account Name": "account_name",

    # 科目类型
    "科目类别": "account_type",
    "科目性质": "account_type",
    "类别": "account_type",
    "Account Type": "account_type",

    # 期初余额
    "期初借方": "opening_debit",
    "期初借方余额": "opening_debit",
    "年初借方": "opening_debit",
    "Beginning Debit": "opening_debit",
    "Opening Debit": "opening_debit",

    "期初贷方": "opening_credit",
    "期初贷方余额": "opening_credit",
    "年初贷方": "opening_credit",
    "Beginning Credit": "opening_credit",
    "Opening Credit": "opening_credit",

    "期初余额": "opening_balance",
    "年初余额": "opening_balance",
    "Beginning Balance": "opening_balance",

    # 本期发生
    "本期借方": "period_debit",
    "本期借方发生额": "period_debit",
    "借方发生额": "period_debit",
    "Debit": "period_debit",
    "Period Debit": "period_debit",

    "本期贷方": "period_credit",
    "本期贷方发生额": "period_credit",
    "贷方发生额": "period_credit",
    "Credit": "period_credit",
    "Period Credit": "period_credit",

    # 期末余额
    "期末借方": "closing_debit",
    "期末借方余额": "closing_debit",
    "Ending Debit": "closing_debit",
    "Closing Debit": "closing_debit",

    "期末贷方": "closing_credit",
    "期末贷方余额": "closing_credit",
    "Ending Credit": "closing_credit",
    "Closing Credit": "closing_credit",

    "期末余额": "closing_balance",
    "余额": "closing_balance",
    "Ending Balance": "closing_balance",
    "Balance": "closing_balance",

    # 方向
    "方向": "direction",
    "余额方向": "direction",
    "借贷方向": "direction",

    # 预算相关
    "预算金额": "budget",
    "预算数": "budget",
    "Budget": "budget",
    "Budgeted": "budget",

    "实际金额": "actual",
    "实际数": "actual",
    "Actual": "actual",

    "上年同期": "prior_year",
    "同期数": "prior_year",
    "Prior Year": "prior_year",
    "Last Year": "prior_year",

    # 凭证相关
    "凭证号": "voucher_no",
    "凭证编号": "voucher_no",
    "Voucher No": "voucher_no",
    "Document Number": "voucher_no",

    "凭证日期": "voucher_date",
    "记账日期": "voucher_date",
    "日期": "date",
    "Date": "date",

    "摘要": "description",
    "摘要说明": "description",
    "业务说明": "description",
    "Description": "description",
    "Memo": "description",

    # 账龄相关
    "0-30天": "0_30",
    "1-30天": "0_30",
    "30天以内": "0_30",
    "0-30 Days": "0_30",

    "31-60天": "31_60",
    "30-60天": "31_60",
    "31-60 Days": "31_60",

    "61-90天": "61_90",
    "60-90天": "61_90",
    "61-90 Days": "61_90",

    "91-180天": "91_180",
    "90-180天": "91_180",
    "91-180 Days": "91_180",

    "181-365天": "181_365",
    "180-365天": "181_365",
    "181-365 Days": "181_365",

    "1年以上": "over_365",
    "365天以上": "over_365",
    "Over 1 Year": "over_365",
    "Over 365 Days": "over_365",

    # 客户/供应商
    "客户名称": "customer_name",
    "客户": "customer_name",
    "Customer": "customer_name",

    "供应商名称": "supplier_name",
    "供应商": "supplier_name",
    "Supplier": "supplier_name",
    "Vendor": "supplier_name",

    # 金额
    "金额": "amount",
    "Amount": "amount",
    "本币金额": "amount",
}

# 科目类型识别（基于中国企业会计准则）
ACCOUNT_TYPE_PATTERNS = {
    "asset": [r"^1\d{3}", r"^资产"],
    "liability": [r"^2\d{3}", r"^负债"],
    "equity": [r"^3\d{3}", r"^权益", r"^所有者权益"],
    "cost": [r"^4\d{3}", r"^成本"],
    "revenue": [r"^5\d{3}", r"^6001", r"^收入", r"^营业收入"],
    "expense": [r"^5\d{3}", r"^6\d{3}", r"^费用", r"^支出"],
}


def detect_account_type(code: str, name: str = "") -> str:
    """
    根据科目编码或名称识别科目类型

    Args:
        code: 科目编码
        name: 科目名称

    Returns:
        科目类型: asset/liability/equity/cost/revenue/expense
    """
    # 优先按编码判断
    if code:
        first_digit = code[0] if code else ""
        type_map = {
            "1": "asset",
            "2": "liability",
            "3": "equity",
            "4": "cost",
            "5": "revenue",  # 损益类（收入）
            "6": "expense",  # 损益类（费用）
        }
        if first_digit in type_map:
            return type_map[first_digit]

    # 按名称判断
    name = name or ""
    if any(kw in name for kw in ["资产", "现金", "银行", "应收", "存货", "固定资产"]):
        return "asset"
    if any(kw in name for kw in ["负债", "应付", "借款", "预收"]):
        return "liability"
    if any(kw in name for kw in ["权益", "股本", "资本", "盈余", "利润"]):
        return "equity"
    if any(kw in name for kw in ["成本"]):
        return "cost"
    if any(kw in name for kw in ["收入", "营业收入"]):
        return "revenue"
    if any(kw in name for kw in ["费用", "支出", "税金"]):
        return "expense"

    return "unknown"


def normalize_column_name(col_name: str) -> str:
    """
    标准化列名

    Args:
        col_name: 原始列名

    Returns:
        标准化后的列名
    """
    if not col_name:
        return ""

    # 保留原始值用于匹配
    original = str(col_name).strip()

    # 清理空白（但保留原始用于映射查找）
    cleaned = re.sub(r'\s+', '', original)

    # 精确匹配（含空格版本）
    if original in COLUMN_MAPPING:
        return COLUMN_MAPPING[original]

    # 精确匹配（去空格版本）
    if cleaned in COLUMN_MAPPING:
        return COLUMN_MAPPING[cleaned]

    # 模糊匹配
    for pattern, std_name in COLUMN_MAPPING.items():
        if pattern in original or original in pattern:
            return std_name
        if pattern in cleaned or cleaned in pattern:
            return std_name

    return cleaned.lower().replace(" ", "_")


def parse_number(value: Any) -> float:
    """
    解析数值，处理各种格式

    Args:
        value: 原始值

    Returns:
        浮点数
    """
    if value is None:
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        # 移除千分位分隔符和货币符号
        value = value.strip()
        value = re.sub(r'[,，\s¥$￥元]', '', value)

        # 处理括号表示的负数 (1000) -> -1000
        if value.startswith('(') and value.endswith(')'):
            value = '-' + value[1:-1]

        # 处理空值
        if value in ['', '-', '--', 'N/A', 'n/a', '无']:
            return 0.0

        try:
            return float(value)
        except ValueError:
            return 0.0

    return 0.0


# ============================================================
# 模板解析器
# ============================================================

def parse_trial_balance(
    rows: List[Dict[str, Any]],
    period: str = "",
    company: str = ""
) -> Dict[str, Any]:
    """
    解析科目余额表

    Args:
        rows: 数据行列表（已标准化列名）
        period: 会计期间
        company: 公司名称

    Returns:
        标准格式的科目余额表 JSON
    """
    accounts = []

    for row in rows:
        code = str(row.get("account_code", "") or "").strip()
        name = str(row.get("account_name", "") or "").strip()

        # 跳过空行或汇总行
        if not code or code in ["合计", "总计", "Total"]:
            continue

        # 解析金额
        opening_debit = parse_number(row.get("opening_debit", 0))
        opening_credit = parse_number(row.get("opening_credit", 0))
        period_debit = parse_number(row.get("period_debit", 0))
        period_credit = parse_number(row.get("period_credit", 0))
        closing_debit = parse_number(row.get("closing_debit", 0))
        closing_credit = parse_number(row.get("closing_credit", 0))

        # 处理合并余额列
        if "opening_balance" in row and opening_debit == 0 and opening_credit == 0:
            bal = parse_number(row.get("opening_balance", 0))
            direction = row.get("direction", "")
            if bal >= 0 or "借" in str(direction):
                opening_debit = abs(bal)
            else:
                opening_credit = abs(bal)

        if "closing_balance" in row and closing_debit == 0 and closing_credit == 0:
            bal = parse_number(row.get("closing_balance", 0))
            direction = row.get("direction", "")
            if bal >= 0 or "借" in str(direction):
                closing_debit = abs(bal)
            else:
                closing_credit = abs(bal)

        # 识别科目类型
        acc_type = row.get("account_type", "")
        if not acc_type or acc_type == "unknown":
            acc_type = detect_account_type(code, name)

        accounts.append({
            "code": code,
            "name": name,
            "type": acc_type,
            "opening_debit": opening_debit,
            "opening_credit": opening_credit,
            "period_debit": period_debit,
            "period_credit": period_credit,
            "closing_debit": closing_debit,
            "closing_credit": closing_credit
        })

    return {
        "template": "trial_balance",
        "period": period,
        "company": company,
        "accounts": accounts,
        "summary": {
            "account_count": len(accounts),
            "total_opening_debit": sum(a["opening_debit"] for a in accounts),
            "total_opening_credit": sum(a["opening_credit"] for a in accounts),
            "total_closing_debit": sum(a["closing_debit"] for a in accounts),
            "total_closing_credit": sum(a["closing_credit"] for a in accounts)
        }
    }


def parse_budget(
    rows: List[Dict[str, Any]],
    period: str = "",
    company: str = ""
) -> Dict[str, Any]:
    """
    解析预算对比表

    Args:
        rows: 数据行列表
        period: 会计期间
        company: 公司名称

    Returns:
        标准格式的预算对比表 JSON
    """
    line_items = []

    for row in rows:
        account = row.get("account_name") or row.get("account_code") or row.get("description", "")
        if not account:
            continue

        actual = parse_number(row.get("actual", 0))
        budget = parse_number(row.get("budget", 0))
        prior_year = parse_number(row.get("prior_year", 0))

        # 计算差异
        variance = actual - budget
        variance_pct = variance / budget if budget != 0 else 0

        yoy_change = actual - prior_year
        yoy_pct = yoy_change / prior_year if prior_year != 0 else 0

        line_items.append({
            "account": str(account).strip(),
            "actual": actual,
            "budget": budget,
            "prior_year": prior_year,
            "variance": round(variance, 2),
            "variance_pct": round(variance_pct, 4),
            "yoy_change": round(yoy_change, 2),
            "yoy_pct": round(yoy_pct, 4)
        })

    return {
        "template": "budget",
        "period": period,
        "company": company,
        "line_items": line_items,
        "summary": {
            "item_count": len(line_items),
            "total_actual": sum(item["actual"] for item in line_items),
            "total_budget": sum(item["budget"] for item in line_items)
        }
    }


def parse_ar_aging(
    rows: List[Dict[str, Any]],
    as_of_date: str = "",
    company: str = ""
) -> Dict[str, Any]:
    """
    解析应收账龄表

    Args:
        rows: 数据行列表
        as_of_date: 截止日期
        company: 公司名称

    Returns:
        标准格式的应收账龄表 JSON
    """
    customers = []
    aging_buckets = ["0_30", "31_60", "61_90", "91_180", "181_365", "over_365"]

    for row in rows:
        customer = row.get("customer_name", "")
        if not customer:
            continue

        aging = {}
        total = 0
        for bucket in aging_buckets:
            amount = parse_number(row.get(bucket, 0))
            aging[bucket] = amount
            total += amount

        customers.append({
            "customer": str(customer).strip(),
            "aging": aging,
            "total": round(total, 2)
        })

    # 汇总
    total_by_bucket = {bucket: sum(c["aging"].get(bucket, 0) for c in customers) for bucket in aging_buckets}

    return {
        "template": "ar_aging",
        "as_of_date": as_of_date,
        "company": company,
        "customers": customers,
        "summary": {
            "customer_count": len(customers),
            "total_receivable": sum(c["total"] for c in customers),
            "by_bucket": total_by_bucket
        }
    }


def parse_ap_aging(
    rows: List[Dict[str, Any]],
    as_of_date: str = "",
    company: str = ""
) -> Dict[str, Any]:
    """
    解析应付账龄表

    Args:
        rows: 数据行列表
        as_of_date: 截止日期
        company: 公司名称

    Returns:
        标准格式的应付账龄表 JSON
    """
    suppliers = []
    aging_buckets = ["0_30", "31_60", "61_90", "91_180", "181_365", "over_365"]

    for row in rows:
        supplier = row.get("supplier_name", "")
        if not supplier:
            continue

        aging = {}
        total = 0
        for bucket in aging_buckets:
            amount = parse_number(row.get(bucket, 0))
            aging[bucket] = amount
            total += amount

        suppliers.append({
            "supplier": str(supplier).strip(),
            "aging": aging,
            "total": round(total, 2)
        })

    # 汇总
    total_by_bucket = {bucket: sum(s["aging"].get(bucket, 0) for s in suppliers) for bucket in aging_buckets}

    return {
        "template": "ap_aging",
        "as_of_date": as_of_date,
        "company": company,
        "suppliers": suppliers,
        "summary": {
            "supplier_count": len(suppliers),
            "total_payable": sum(s["total"] for s in suppliers),
            "by_bucket": total_by_bucket
        }
    }


def parse_voucher(
    rows: List[Dict[str, Any]],
    period: str = "",
    company: str = ""
) -> Dict[str, Any]:
    """
    解析记账凭证

    Args:
        rows: 数据行列表
        period: 会计期间
        company: 公司名称

    Returns:
        标准格式的凭证列表 JSON
    """
    vouchers = {}

    for row in rows:
        voucher_no = row.get("voucher_no", "")
        if not voucher_no:
            continue

        voucher_no = str(voucher_no).strip()

        if voucher_no not in vouchers:
            vouchers[voucher_no] = {
                "voucher_no": voucher_no,
                "date": row.get("voucher_date", row.get("date", "")),
                "entries": [],
                "total_debit": 0,
                "total_credit": 0
            }

        debit = parse_number(row.get("period_debit", row.get("debit", 0)))
        credit = parse_number(row.get("period_credit", row.get("credit", 0)))

        vouchers[voucher_no]["entries"].append({
            "account_code": row.get("account_code", ""),
            "account_name": row.get("account_name", ""),
            "description": row.get("description", ""),
            "debit": debit,
            "credit": credit
        })
        vouchers[voucher_no]["total_debit"] += debit
        vouchers[voucher_no]["total_credit"] += credit

    voucher_list = list(vouchers.values())

    return {
        "template": "voucher",
        "period": period,
        "company": company,
        "vouchers": voucher_list,
        "summary": {
            "voucher_count": len(voucher_list),
            "entry_count": sum(len(v["entries"]) for v in voucher_list)
        }
    }


def parse_cash_flow(
    rows: List[Dict[str, Any]],
    period: str = "",
    company: str = ""
) -> Dict[str, Any]:
    """
    解析现金流数据

    Args:
        rows: 数据行列表
        period: 会计期间
        company: 公司名称

    Returns:
        标准格式的现金流数据 JSON
    """
    items = []

    for row in rows:
        desc = row.get("description") or row.get("account_name", "")
        if not desc:
            continue

        amount = parse_number(row.get("amount", 0))
        date = row.get("date", "")

        items.append({
            "description": str(desc).strip(),
            "amount": amount,
            "date": str(date) if date else "",
            "category": row.get("category", "operating")
        })

    return {
        "template": "cash_flow",
        "period": period,
        "company": company,
        "items": items,
        "summary": {
            "item_count": len(items),
            "total_inflow": sum(i["amount"] for i in items if i["amount"] > 0),
            "total_outflow": sum(i["amount"] for i in items if i["amount"] < 0),
            "net_flow": sum(i["amount"] for i in items)
        }
    }


# ============================================================
# Excel 解析入口
# ============================================================

def read_excel_rows(
    filepath: str,
    sheet_name: Optional[str] = None,
    header_row: int = 1
) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    读取 Excel 文件，返回标准化的数据行

    Args:
        filepath: Excel 文件路径
        sheet_name: 工作表名称（None 为第一个）
        header_row: 表头行号（从1开始）

    Returns:
        (标准化列名列表, 数据行列表)
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("需要安装 openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # 选择工作表
    if sheet_name:
        # 模糊匹配
        target_sheet = None
        for name in wb.sheetnames:
            if sheet_name in name or name in sheet_name:
                target_sheet = wb[name]
                break
        if target_sheet is None:
            target_sheet = wb.active
    else:
        target_sheet = wb.active

    # 读取表头
    header_cells = list(target_sheet.iter_rows(min_row=header_row, max_row=header_row))[0]
    headers = [normalize_column_name(cell.value) for cell in header_cells]

    # 读取数据行
    rows = []
    for row in target_sheet.iter_rows(min_row=header_row + 1):
        row_data = {}
        for i, cell in enumerate(row):
            if i < len(headers) and headers[i]:
                row_data[headers[i]] = cell.value
        if any(v is not None for v in row_data.values()):
            rows.append(row_data)

    wb.close()
    return headers, rows


def read_csv_rows(
    filepath: str,
    encoding: str = "utf-8",
    header_row: int = 1
) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    读取 CSV 文件，返回标准化的数据行

    Args:
        filepath: CSV 文件路径
        encoding: 文件编码
        header_row: 表头行号（从1开始）

    Returns:
        (标准化列名列表, 数据行列表)
    """
    import csv

    # 尝试不同编码
    encodings = [encoding, "utf-8", "gbk", "gb2312", "utf-8-sig"]

    for enc in encodings:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                reader = csv.reader(f)
                all_rows = list(reader)
                break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError(f"无法解码文件: {filepath}")

    # 读取表头
    if header_row > len(all_rows):
        return [], []

    headers = [normalize_column_name(h) for h in all_rows[header_row - 1]]

    # 读取数据行
    rows = []
    for row in all_rows[header_row:]:
        row_data = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                row_data[headers[i]] = value
        if any(v for v in row_data.values()):
            rows.append(row_data)

    return headers, rows


def parse_erp_file(
    filepath: str,
    template: str,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    period: str = "",
    company: str = "",
    encoding: str = "utf-8"
) -> Dict[str, Any]:
    """
    解析 ERP 导出文件

    Args:
        filepath: 文件路径（Excel 或 CSV）
        template: 模板类型
            - trial_balance: 科目余额表
            - budget: 预算对比表
            - ar_aging: 应收账龄表
            - ap_aging: 应付账龄表
            - voucher: 记账凭证
            - cash: 现金流数据
        sheet_name: 工作表名称（仅 Excel）
        header_row: 表头行号
        period: 会计期间
        company: 公司名称
        encoding: 文件编码（仅 CSV）

    Returns:
        解析后的标准 JSON 数据
    """
    filepath = str(filepath)

    # 判断文件类型
    if filepath.endswith(('.xlsx', '.xls')):
        headers, rows = read_excel_rows(filepath, sheet_name, header_row)
    elif filepath.endswith('.csv'):
        headers, rows = read_csv_rows(filepath, encoding, header_row)
    else:
        raise ValueError(f"不支持的文件格式: {filepath}")

    # 根据模板解析
    parsers = {
        "trial_balance": parse_trial_balance,
        "tb": parse_trial_balance,
        "budget": parse_budget,
        "variance": parse_budget,
        "ar_aging": parse_ar_aging,
        "ar": parse_ar_aging,
        "ap_aging": parse_ap_aging,
        "ap": parse_ap_aging,
        "voucher": parse_voucher,
        "cash": parse_cash_flow,
        "cash_flow": parse_cash_flow
    }

    parser = parsers.get(template.lower())
    if not parser:
        raise ValueError(f"不支持的模板类型: {template}。支持: {list(parsers.keys())}")

    if template in ["ar_aging", "ar", "ap_aging", "ap"]:
        return parser(rows, as_of_date=period, company=company)
    else:
        return parser(rows, period=period, company=company)


# ============================================================
# 工具注册
# ============================================================

ERP_PARSER_TOOLS = {
    "parse_erp_file": {
        "function": parse_erp_file,
        "description": "解析金蝶/用友/SAP导出的Excel/CSV文件",
        "parameters": ["filepath", "template", "sheet_name", "header_row", "period", "company"]
    },
    "parse_trial_balance": {
        "function": parse_trial_balance,
        "description": "解析科目余额表数据",
        "parameters": ["rows", "period", "company"]
    },
    "parse_budget": {
        "function": parse_budget,
        "description": "解析预算对比表数据",
        "parameters": ["rows", "period", "company"]
    },
    "parse_ar_aging": {
        "function": parse_ar_aging,
        "description": "解析应收账龄表数据",
        "parameters": ["rows", "as_of_date", "company"]
    },
    "parse_ap_aging": {
        "function": parse_ap_aging,
        "description": "解析应付账龄表数据",
        "parameters": ["rows", "as_of_date", "company"]
    }
}
