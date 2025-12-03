# -*- coding: utf-8 -*-
"""
M&A 模型（并购分析）

公司并购交易分析工具

本模块提供两种使用方式:

1. 原子工具（推荐LLM使用）:
   from fin_tools.tools import calc_offer_price, calc_accretion_dilution, ...
   每个工具独立运行，可自由组合

2. MAModel类（封装接口）:
   ma = MAModel()
   result = ma.build(inputs)
   内部调用原子工具，提供统一接口
"""

from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from ..core import ModelResult, FinancialModel
from ..tools import ma_tools


class MAModel(FinancialModel):
    """
    M&A估值工具集

    使用方法:
        ma = MAModel()
        result = ma.build(inputs)

    或者使用单独的分析方法:
        ma = MAModel()
        offer = ma.calc_offer_price(target_price=50, shares=200_000_000, premium=0.30)
        ad = ma.analyze_accretion_dilution(...)
    """

    def __init__(self, name: str = "M&A模型"):
        super().__init__(name=name, scenario="ma_analysis")

    # ==================== 交易结构计算 ====================

    def calc_offer_price(self,
                         target_share_price: float,
                         target_shares: float,
                         premium_percent: float = 0.0) -> ModelResult:
        """
        计算收购报价

        公式: Offer Price = Target Share Price × (1 + Premium)

        Args:
            target_share_price: 目标公司当前股价
            target_shares: 目标公司流通股数
            premium_percent: 收购溢价比例

        Returns:
            ModelResult
        """
        result = ma_tools.calc_offer_price(
            target_share_price=target_share_price,
            target_shares=target_shares,
            premium_percent=premium_percent
        )

        return ModelResult(
            value=result["total_purchase_price"],
            formula=result["formula"],
            inputs=result["inputs"]
        )

    def calc_funding_mix(self,
                         total_purchase_price: float,
                         cash_percent: float,
                         stock_percent: float,
                         acquirer_share_price: float,
                         new_debt: float = 0) -> ModelResult:
        """
        计算融资结构

        Args:
            total_purchase_price: 总收购价格
            cash_percent: 现金支付比例
            stock_percent: 股票支付比例
            acquirer_share_price: 收购方股价
            new_debt: 新增债务融资

        Returns:
            ModelResult
        """
        result = ma_tools.calc_funding_mix(
            total_purchase_price=total_purchase_price,
            cash_percent=cash_percent,
            stock_percent=stock_percent,
            acquirer_share_price=acquirer_share_price,
            new_debt=new_debt
        )

        return ModelResult(
            value=result["new_shares_issued"],
            formula=result["formula"],
            inputs={
                "funding_breakdown": result["funding_breakdown"],
                "cash_amount": result["cash_amount"],
                "stock_amount": result["stock_amount"],
                "new_shares_issued": result["new_shares_issued"],
                "new_debt": result["new_debt"]
            }
        )

    def calc_goodwill(self,
                      purchase_price: float,
                      target_book_value: float,
                      fair_value_adjustments: Optional[Dict[str, float]] = None) -> ModelResult:
        """
        计算商誉

        公式: Goodwill = Purchase Price - Adjusted Net Assets

        Args:
            purchase_price: 收购价格
            target_book_value: 目标公司账面净资产
            fair_value_adjustments: 公允价值调整

        Returns:
            ModelResult
        """
        result = ma_tools.calc_goodwill(
            purchase_price=purchase_price,
            target_book_value=target_book_value,
            fair_value_adjustments=fair_value_adjustments
        )

        return ModelResult(
            value=result["goodwill"],
            formula=result["formula"],
            inputs={
                "purchase_price": purchase_price,
                "target_book_value": target_book_value,
                "adjusted_net_assets": result["adjusted_net_assets"],
                "is_bargain_purchase": result["is_bargain_purchase"]
            }
        )

    # ==================== 增厚/稀释分析 ====================

    def analyze_accretion_dilution(self,
                                   acquirer_net_income: float,
                                   acquirer_shares: float,
                                   target_net_income: float,
                                   purchase_price: float,
                                   cash_percent: float,
                                   stock_percent: float,
                                   acquirer_share_price: float,
                                   cost_of_debt: float = 0.05,
                                   tax_rate: float = 0.25,
                                   foregone_interest_rate: float = 0.03,
                                   synergies: float = 0,
                                   intangible_amort: float = 0) -> ModelResult:
        """
        增厚/稀释分析

        Args:
            acquirer_net_income: 收购方净利润
            acquirer_shares: 收购方流通股数
            target_net_income: 目标公司净利润
            purchase_price: 收购价格
            cash_percent: 现金支付比例
            stock_percent: 股票支付比例
            acquirer_share_price: 收购方股价
            cost_of_debt: 新债务成本
            tax_rate: 税率
            foregone_interest_rate: 损失的现金利息收益率
            synergies: 协同效应（税前）
            intangible_amort: 无形资产年摊销额

        Returns:
            ModelResult
        """
        result = ma_tools.calc_accretion_dilution(
            acquirer_net_income=acquirer_net_income,
            acquirer_shares=acquirer_shares,
            target_net_income=target_net_income,
            purchase_price=purchase_price,
            cash_percent=cash_percent,
            stock_percent=stock_percent,
            acquirer_share_price=acquirer_share_price,
            cost_of_debt=cost_of_debt,
            tax_rate=tax_rate,
            foregone_interest_rate=foregone_interest_rate,
            synergies=synergies,
            intangible_amort=intangible_amort
        )

        return ModelResult(
            value=result["accretion_dilution"]["percent"],
            formula=result["formula"],
            inputs={
                "standalone": result["standalone"],
                "pro_forma": result["pro_forma"],
                "accretion_dilution": result["accretion_dilution"]
            }
        )

    # ==================== 协同效应分析 ====================

    def analyze_synergies(self,
                          cost_synergies,
                          revenue_synergies=0,
                          integration_costs=0,
                          years: int = 5,
                          discount_rate: float = 0.10,
                          tax_rate: float = 0.25,
                          margin_on_revenue: float = 0.20) -> ModelResult:
        """
        协同效应分析

        Args:
            cost_synergies: 成本协同（税前）
            revenue_synergies: 收入协同
            integration_costs: 整合成本
            years: 分析年数
            discount_rate: 折现率
            tax_rate: 税率
            margin_on_revenue: 收入协同的利润率

        Returns:
            ModelResult
        """
        result = ma_tools.calc_synergies(
            cost_synergies=cost_synergies,
            revenue_synergies=revenue_synergies,
            integration_costs=integration_costs,
            years=years,
            discount_rate=discount_rate,
            tax_rate=tax_rate,
            margin_on_revenue=margin_on_revenue
        )

        return ModelResult(
            value=result["total_synergies_pv_with_terminal"],
            formula=result["formula"],
            inputs={
                "yearly_details": result["yearly_details"],
                "terminal_value": result["terminal_value"],
                "summary": result["summary"]
            }
        )

    # ==================== 盈亏平衡分析 ====================

    def calc_breakeven_synergies(self,
                                 acquirer_net_income: float,
                                 acquirer_shares: float,
                                 target_net_income: float,
                                 purchase_price: float,
                                 cash_percent: float,
                                 stock_percent: float,
                                 acquirer_share_price: float,
                                 cost_of_debt: float = 0.05,
                                 tax_rate: float = 0.25,
                                 foregone_interest_rate: float = 0.03,
                                 intangible_amort: float = 0) -> ModelResult:
        """
        计算盈亏平衡所需协同效应

        Args:
            （同 analyze_accretion_dilution）

        Returns:
            ModelResult: 使EPS不稀释所需的年化协同效应（税前）
        """
        result = ma_tools.calc_breakeven(
            acquirer_net_income=acquirer_net_income,
            acquirer_shares=acquirer_shares,
            target_net_income=target_net_income,
            purchase_price=purchase_price,
            cash_percent=cash_percent,
            stock_percent=stock_percent,
            acquirer_share_price=acquirer_share_price,
            cost_of_debt=cost_of_debt,
            tax_rate=tax_rate,
            foregone_interest_rate=foregone_interest_rate,
            intangible_amort=intangible_amort
        )

        return ModelResult(
            value=result["synergies_needed"],
            formula=result["formula"],
            inputs={
                "current_dilution": result["current_dilution"],
                "analysis": result["analysis"],
                "interpretation": result["interpretation"]
            }
        )

    # ==================== 敏感性分析 ====================

    def sensitivity_payment_mix(self,
                                base_inputs: dict,
                                cash_range: List[float] = None) -> Dict[str, Any]:
        """
        支付方式敏感性分析

        Args:
            base_inputs: 基础输入参数
            cash_range: 现金比例范围 [0.2, 0.4, 0.6, 0.8, 1.0]

        Returns:
            dict: EPS增厚/稀释敏感性矩阵
        """
        if cash_range is None:
            cash_range = [0.0, 0.25, 0.50, 0.75, 1.0]

        results = []

        for cash_pct in cash_range:
            stock_pct = 1.0 - cash_pct

            inputs = base_inputs.copy()
            inputs["deal_terms"] = inputs.get("deal_terms", {}).copy()
            inputs["deal_terms"]["cash_percent"] = cash_pct
            inputs["deal_terms"]["stock_percent"] = stock_pct

            result = self.build(inputs)
            ad = result["accretion_dilution"]["without_synergies"]["accretion_dilution"]

            results.append({
                "cash_percent": f"{cash_pct:.0%}",
                "stock_percent": f"{stock_pct:.0%}",
                "eps_change": f"{ad['percent']:.2%}",
                "status": ad["status"]
            })

        return {
            "headers": {"metric": "EPS Accretion/Dilution"},
            "data": results
        }

    def sensitivity_premium(self,
                            base_inputs: dict,
                            premium_range: List[float] = None) -> Dict[str, Any]:
        """
        溢价敏感性分析

        Args:
            base_inputs: 基础输入参数
            premium_range: 溢价范围 [0.1, 0.2, 0.3, 0.4, 0.5]

        Returns:
            dict: EPS增厚/稀释敏感性矩阵
        """
        if premium_range is None:
            premium_range = [0.10, 0.20, 0.30, 0.40, 0.50]

        results = []

        for premium in premium_range:
            inputs = base_inputs.copy()
            inputs["deal_terms"] = inputs.get("deal_terms", {}).copy()
            inputs["deal_terms"]["premium_percent"] = premium

            result = self.build(inputs)
            ad = result["accretion_dilution"]["without_synergies"]["accretion_dilution"]

            results.append({
                "premium": f"{premium:.0%}",
                "purchase_price": result["deal_summary"]["purchase_price"],
                "eps_change": f"{ad['percent']:.2%}",
                "status": ad["status"],
                "breakeven_synergies": result["breakeven"]["synergies_needed"]
            })

        return {
            "headers": {"metric": "Premium Sensitivity"},
            "data": results
        }

    # ==================== 主构建方法 ====================

    def build(self, inputs: dict) -> Dict[str, Any]:
        """
        构建完整M&A模型

        内部调用原子工具实现，保持接口兼容。

        Args:
            inputs: M&A输入参数，包含：
                - acquirer: 收购方财务数据
                    - share_price: 股价
                    - shares_outstanding: 流通股数
                    - net_income: 净利润
                    - total_assets: 总资产
                    - total_liabilities: 总负债
                    - total_equity: 总权益
                    - cash: 现金
                    - goodwill: 商誉
                - target: 目标公司财务数据（同上）
                - deal_terms: 交易条款
                    - premium_percent: 溢价比例
                    - cash_percent: 现金支付比例
                    - stock_percent: 股票支付比例
                - ppa: 购买价格分摊（可选）
                - synergies: 协同效应假设（可选）
                - assumptions: 其他假设

        Returns:
            dict: 完整M&A分析结果
        """
        # 使用原子工具快捷构建
        result = ma_tools.ma_quick_build(inputs)

        # 添加模型名称
        result["_meta"]["model_name"] = self.name

        # 转换为与类方法兼容的格式
        def to_model_result_dict(d: dict) -> dict:
            """转换为 ModelResult.to_dict() 格式"""
            if d is None:
                return None
            return {
                "value": d.get("value") or d.get("total_purchase_price") or d.get("goodwill"),
                "formula": d.get("formula"),
                "inputs": d.get("inputs", {})
            }

        return result

    # ==================== Excel导出 ====================

    def to_excel(self, result: Dict[str, Any], filepath: str) -> str:
        """
        导出M&A分析结果到Excel

        Args:
            result: build() 返回的结果
            filepath: 输出文件路径

        Returns:
            str: 文件路径
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")

        wb = openpyxl.Workbook()

        # ===== 交易摘要 =====
        ws = wb.active
        ws.title = "交易摘要"

        # 标题样式
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

        # 交易摘要
        ws["A1"] = "M&A 交易分析"
        ws["A1"].font = title_font

        row = 3
        ws[f"A{row}"] = "交易条款"
        ws[f"A{row}"].font = header_font

        deal = result["deal_summary"]
        row += 1
        ws[f"A{row}"] = "收购价格"
        ws[f"B{row}"] = deal["purchase_price"]
        ws[f"B{row}"].number_format = "#,##0"

        row += 1
        ws[f"A{row}"] = "溢价比例"
        ws[f"B{row}"] = deal["premium"]["percent"]
        ws[f"B{row}"].number_format = "0.0%"

        row += 1
        ws[f"A{row}"] = "溢价金额"
        ws[f"B{row}"] = deal["premium"]["value"]
        ws[f"B{row}"].number_format = "#,##0"

        funding = deal["funding"]
        row += 2
        ws[f"A{row}"] = "融资结构"
        ws[f"A{row}"].font = header_font

        row += 1
        ws[f"A{row}"] = "现金支付"
        ws[f"B{row}"] = funding["cash_amount"]
        ws[f"B{row}"].number_format = "#,##0"
        ws[f"C{row}"] = funding["funding_breakdown"]["cash"]["percent"]
        ws[f"C{row}"].number_format = "0.0%"

        row += 1
        ws[f"A{row}"] = "股票支付"
        ws[f"B{row}"] = funding["stock_amount"]
        ws[f"B{row}"].number_format = "#,##0"
        ws[f"C{row}"] = funding["funding_breakdown"]["stock"]["percent"]
        ws[f"C{row}"].number_format = "0.0%"

        row += 1
        ws[f"A{row}"] = "新发行股数"
        ws[f"B{row}"] = funding["new_shares_issued"]
        ws[f"B{row}"].number_format = "#,##0"

        # ===== 增厚/稀释分析 =====
        row += 2
        ws[f"A{row}"] = "增厚/稀释分析"
        ws[f"A{row}"].font = header_font

        ad = result["accretion_dilution"]["without_synergies"]
        row += 1
        ws[f"A{row}"] = "收购前EPS"
        ws[f"B{row}"] = ad["standalone"]["eps"]
        ws[f"B{row}"].number_format = "#,##0.00"

        row += 1
        ws[f"A{row}"] = "合并后EPS"
        ws[f"B{row}"] = ad["pro_forma"]["eps"]
        ws[f"B{row}"].number_format = "#,##0.00"

        row += 1
        ws[f"A{row}"] = "EPS变化"
        ws[f"B{row}"] = ad["accretion_dilution"]["percent"]
        ws[f"B{row}"].number_format = "0.00%"
        ws[f"C{row}"] = ad["accretion_dilution"]["status"]

        # 盈亏平衡
        row += 2
        ws[f"A{row}"] = "盈亏平衡分析"
        ws[f"A{row}"].font = header_font

        be = result["breakeven"]
        row += 1
        ws[f"A{row}"] = "所需年化协同效应"
        ws[f"B{row}"] = be["synergies_needed"]
        ws[f"B{row}"].number_format = "#,##0"

        # 商誉
        row += 2
        ws[f"A{row}"] = "购买价格分摊"
        ws[f"A{row}"].font = header_font

        ppa = result["purchase_price_allocation"]
        row += 1
        ws[f"A{row}"] = "商誉"
        ws[f"B{row}"] = ppa["goodwill"]
        ws[f"B{row}"].number_format = "#,##0"

        row += 1
        ws[f"A{row}"] = "调整后净资产"
        ws[f"B{row}"] = ppa["adjusted_net_assets"]
        ws[f"B{row}"].number_format = "#,##0"

        if ppa["is_bargain_purchase"]:
            row += 1
            ws[f"A{row}"] = "注: 廉价收购（负商誉）"

        # 协同效应
        if result.get("synergies"):
            row += 2
            ws[f"A{row}"] = "协同效应分析"
            ws[f"A{row}"].font = header_font

            syn = result["synergies"]
            row += 1
            ws[f"A{row}"] = "协同效应现值（含终值）"
            ws[f"B{row}"] = syn["total_synergies_pv_with_terminal"]
            ws[f"B{row}"].number_format = "#,##0"

            if result.get("synergy_coverage"):
                row += 1
                ws[f"A{row}"] = "协同效应覆盖率"
                ws[f"B{row}"] = result["synergy_coverage"]["coverage_ratio"]
                ws[f"B{row}"].number_format = "0.00x"

        # 调整列宽
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 12

        # 保存
        wb.save(filepath)
        return filepath
