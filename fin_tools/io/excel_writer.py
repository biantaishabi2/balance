# -*- coding: utf-8 -*-
"""
Excel 导出工具

将建模结果导出为格式化的 Excel 文件，方便人类阅读
"""

from typing import Dict, Any, List
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelWriter:
    """
    Excel 导出工具

    将建模结果导出为格式化的 Excel 文件

    使用方法:
        writer = ExcelWriter()
        writer.write_three_statement(result, "三表模型")
        writer.write_dcf(dcf_result, "DCF估值")
        writer.save("output.xlsx")
    """

    def __init__(self):
        if not HAS_OPENPYXL:
            raise ImportError("请安装 openpyxl: pip install openpyxl")

        self.wb = Workbook()
        # 删除默认的 sheet
        self.wb.remove(self.wb.active)

        # 样式定义
        self.title_font = Font(bold=True, size=14)
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        self.number_font = Font(size=11)
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _format_number(self, value, decimals=2):
        """格式化数字"""
        if isinstance(value, (int, float)):
            if abs(value) >= 10000:
                return f"{value:,.{decimals}f}"
            else:
                return f"{value:.{decimals}f}"
        return str(value)

    def _set_column_width(self, ws, col, width):
        """设置列宽"""
        ws.column_dimensions[get_column_letter(col)].width = width

    def _write_title(self, ws, row, title):
        """写标题"""
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = self.title_font
        return row + 1

    def _write_header_row(self, ws, row, headers, start_col=1):
        """写表头行"""
        for i, header in enumerate(headers):
            cell = ws.cell(row=row, column=start_col + i, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center')
        return row + 1

    def _write_data_row(self, ws, row, data, start_col=1, is_total=False):
        """写数据行"""
        for i, value in enumerate(data):
            cell = ws.cell(row=row, column=start_col + i, value=value)
            cell.border = self.thin_border
            if i == 0:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='right')
            if is_total:
                cell.font = Font(bold=True)
        return row + 1

    def write_three_statement(self, result: Dict[str, Any], sheet_name: str = "三表模型"):
        """
        写入三表模型结果

        Args:
            result: ThreeStatementModel.build() 的返回结果
            sheet_name: 工作表名称
        """
        ws = self.wb.create_sheet(title=sheet_name)

        # 设置列宽
        self._set_column_width(ws, 1, 20)
        self._set_column_width(ws, 2, 18)
        self._set_column_width(ws, 3, 40)

        row = 1

        # 元信息
        meta = result.get("_meta", {})
        if meta:
            row = self._write_title(ws, row, f"{meta.get('company', '')} 财务模型 - {meta.get('scenario', '')}")
            row += 1

        # ===== 利润表 =====
        row = self._write_title(ws, row, "利润表")
        row = self._write_header_row(ws, row, ["科目", "金额(万元)", "公式"])

        income = result.get("income_statement", {})
        income_items = [
            ("revenue", "营业收入"),
            ("cost", "营业成本"),
            ("gross_profit", "毛利"),
            ("opex", "营业费用"),
            ("ebit", "营业利润"),
            ("interest", "利息费用"),
            ("ebt", "税前利润"),
            ("tax", "所得税"),
            ("net_income", "净利润"),
        ]

        for key, label in income_items:
            if key in income:
                item = income[key]
                value = item.get("value", 0)
                formula = item.get("formula", "")
                is_total = key in ["gross_profit", "ebit", "ebt", "net_income"]
                row = self._write_data_row(ws, row, [label, value, formula], is_total=is_total)

        row += 1

        # ===== 资产负债表 =====
        row = self._write_title(ws, row, "资产负债表")
        row = self._write_header_row(ws, row, ["科目", "金额(万元)", "说明"])

        bs = result.get("balance_sheet", {})

        # 资产
        ws.cell(row=row, column=1, value="【资产】").font = Font(bold=True)
        row += 1

        assets = bs.get("assets", {})
        for key, label in [("cash", "现金"), ("receivable", "应收账款"),
                           ("inventory", "存货"), ("fixed_assets_net", "固定资产净值")]:
            if key in assets:
                item = assets[key]
                value = item.get("value", 0) if isinstance(item, dict) else item
                row = self._write_data_row(ws, row, [f"  {label}", value, ""])

        if "total_assets" in assets:
            row = self._write_data_row(ws, row, ["资产合计", assets["total_assets"].get("value", 0), ""],
                                       is_total=True)

        # 负债
        ws.cell(row=row, column=1, value="【负债】").font = Font(bold=True)
        row += 1

        liabilities = bs.get("liabilities", {})
        for key, label in [("payable", "应付账款"), ("debt", "有息负债")]:
            if key in liabilities:
                item = liabilities[key]
                value = item.get("value", 0) if isinstance(item, dict) else item
                row = self._write_data_row(ws, row, [f"  {label}", value, ""])

        if "total_liabilities" in liabilities:
            row = self._write_data_row(ws, row, ["负债合计", liabilities["total_liabilities"].get("value", 0), ""],
                                       is_total=True)

        # 权益
        ws.cell(row=row, column=1, value="【权益】").font = Font(bold=True)
        row += 1

        equity = bs.get("equity", {})
        if "total_equity" in equity:
            row = self._write_data_row(ws, row, ["权益合计", equity["total_equity"].get("value", 0), ""],
                                       is_total=True)

        row += 1

        # ===== 现金流量表 =====
        row = self._write_title(ws, row, "现金流量表")
        row = self._write_header_row(ws, row, ["科目", "金额(万元)", "公式"])

        cf = result.get("cash_flow", {})
        cf_items = [
            ("operating", "经营活动现金流"),
            ("investing", "投资活动现金流"),
            ("financing", "筹资活动现金流"),
            ("net_change", "现金净变动"),
            ("closing_cash", "期末现金"),
        ]

        for key, label in cf_items:
            if key in cf:
                item = cf[key]
                value = item.get("value", 0)
                formula = item.get("formula", "")
                is_total = key in ["net_change", "closing_cash"]
                row = self._write_data_row(ws, row, [label, value, formula], is_total=is_total)

        row += 1

        # ===== 配平检验 =====
        validation = result.get("validation", {})
        if validation:
            row = self._write_title(ws, row, "配平检验")
            status = "✓ 配平" if validation.get("is_balanced") else "✗ 不配平"
            diff = validation.get("difference", 0)
            row = self._write_data_row(ws, row, ["状态", status, f"差额: {diff:.2f}"])

    def write_dcf(self, result: Dict[str, Any], sheet_name: str = "DCF估值"):
        """
        写入 DCF 估值结果

        Args:
            result: DCFModel.valuate() 的返回结果
            sheet_name: 工作表名称
        """
        ws = self.wb.create_sheet(title=sheet_name)

        self._set_column_width(ws, 1, 20)
        self._set_column_width(ws, 2, 18)
        self._set_column_width(ws, 3, 40)

        row = 1
        row = self._write_title(ws, row, "DCF 估值模型")
        row += 1

        # 折现率
        row = self._write_title(ws, row, "折现率计算")
        row = self._write_header_row(ws, row, ["项目", "值", "公式"])

        if "cost_of_equity" in result:
            re = result["cost_of_equity"]
            row = self._write_data_row(ws, row, ["股权成本 (Re)", f"{re['value']:.2%}", re.get("formula", "")])

        if "wacc" in result:
            wacc = result["wacc"]
            row = self._write_data_row(ws, row, ["WACC", f"{wacc['value']:.2%}", wacc.get("formula", "")],
                                       is_total=True)

        row += 1

        # FCF 预测
        if "_inputs" in result and "fcf_projections" in result["_inputs"]:
            row = self._write_title(ws, row, "自由现金流预测")
            row = self._write_header_row(ws, row, ["年份", "FCF (万元)", ""])

            for year, fcf in result["_inputs"]["fcf_projections"].items():
                row = self._write_data_row(ws, row, [year, fcf, ""])

            row += 1

        # 估值结果
        row = self._write_title(ws, row, "估值结果")
        row = self._write_header_row(ws, row, ["项目", "值 (万元)", "公式"])

        items = [
            ("terminal_value", "终值"),
            ("enterprise_value", "企业价值"),
            ("equity_value", "股权价值"),
        ]

        for key, label in items:
            if key in result:
                item = result[key]
                row = self._write_data_row(ws, row, [label, item["value"], item.get("formula", "")])

        if "per_share_value" in result:
            psv = result["per_share_value"]
            row = self._write_data_row(ws, row, ["每股价值 (元)", psv["value"], psv.get("formula", "")],
                                       is_total=True)

        row += 1

        # 敏感性分析
        if "sensitivity_matrix" in result and result["sensitivity_matrix"]:
            sens = result["sensitivity_matrix"]
            row = self._write_title(ws, row, "敏感性分析 (每股价值)")

            # 获取列头
            if sens.get("data"):
                first_row = sens["data"][0]
                headers = ["WACC \\ g"] + [k for k in first_row.keys() if k != "wacc"]
                row = self._write_header_row(ws, row, headers)

                for data_row in sens["data"]:
                    values = [data_row["wacc"]] + [data_row.get(k, "N/A") for k in headers[1:]]
                    row = self._write_data_row(ws, row, values)

    def write_scenario_comparison(self, result: Dict[str, Any], sheet_name: str = "场景对比"):
        """
        写入场景对比结果

        Args:
            result: ScenarioManager.compare_scenarios() 的返回结果
            sheet_name: 工作表名称
        """
        ws = self.wb.create_sheet(title=sheet_name)

        self._set_column_width(ws, 1, 15)
        for i in range(2, 6):
            self._set_column_width(ws, i, 18)

        row = 1
        row = self._write_title(ws, row, "场景对比分析")
        row += 1

        comparison = result.get("scenario_comparison", result)

        if "headers" in comparison:
            headers = comparison["headers"]
            row = self._write_header_row(ws, row, headers)

            for data_row in comparison.get("rows", []):
                values = [data_row.get("metric", "")]
                for h in headers[1:]:
                    values.append(data_row.get(h, ""))
                row = self._write_data_row(ws, row, values)

        # 验证
        if "validation" in comparison:
            row += 1
            v = comparison["validation"]
            status = "✓ 通过" if v.get("is_valid") else "✗ 不通过"
            row = self._write_data_row(ws, row, ["场景排序验证", status, v.get("message", "")])

    def write_sensitivity_2d(self, result: Dict[str, Any], sheet_name: str = "敏感性分析"):
        """
        写入双参数敏感性矩阵

        Args:
            result: ScenarioManager.sensitivity_2d() 的返回结果
            sheet_name: 工作表名称
        """
        ws = self.wb.create_sheet(title=sheet_name)

        row = 1

        headers = result.get("headers", {})
        param1 = headers.get("rows", "参数1")
        param2 = headers.get("columns", "参数2")
        metric = result.get("output_metric", "")

        row = self._write_title(ws, row, f"敏感性分析: {metric}")
        row += 1

        if result.get("data"):
            first_row = result["data"][0]
            col_headers = [f"{param1} \\ {param2}"] + [k for k in first_row.keys() if k != param1]
            row = self._write_header_row(ws, row, col_headers)

            for data_row in result["data"]:
                p1_val = data_row.get(param1, "")
                if isinstance(p1_val, float) and p1_val < 1:
                    p1_val = f"{p1_val:.1%}"
                values = [p1_val] + [data_row.get(k, "") for k in col_headers[1:]]
                row = self._write_data_row(ws, row, values)

    def write_advanced_features(self, deferred_tax=None, impairment=None, lease=None,
                                 sheet_name: str = "高级功能"):
        """
        写入高级功能结果

        Args:
            deferred_tax: 递延税结果
            impairment: 减值测试结果
            lease: 租赁资本化结果
            sheet_name: 工作表名称
        """
        ws = self.wb.create_sheet(title=sheet_name)

        self._set_column_width(ws, 1, 25)
        self._set_column_width(ws, 2, 18)
        self._set_column_width(ws, 3, 35)

        row = 1

        # ===== 递延税 =====
        if deferred_tax:
            row = self._write_title(ws, row, "递延税计算")
            row = self._write_header_row(ws, row, ["项目", "值", "说明"])

            for scenario, data in deferred_tax.items():
                ws.cell(row=row, column=1, value=f"【{scenario}】").font = Font(bold=True)
                row += 1

                items = [
                    ("current_tax", "当期税"),
                    ("dta_change", "DTA变动"),
                    ("closing_dta", "期末DTA"),
                    ("tax_expense", "所得税费用"),
                ]
                for key, label in items:
                    if key in data:
                        item = data[key]
                        value = item.get("value", 0)
                        formula = item.get("formula", "")
                        row = self._write_data_row(ws, row, [f"  {label}", value, formula])

            row += 1

        # ===== 减值测试 =====
        if impairment:
            row = self._write_title(ws, row, "资产减值测试")
            row = self._write_header_row(ws, row, ["项目", "值", "公式"])

            test = impairment.get("impairment_test", impairment)

            items = [
                ("carrying_value", "账面价值"),
                ("recoverable_amount", "可回收金额"),
                ("impairment_loss", "减值损失"),
                ("new_carrying_value", "新账面价值"),
            ]

            for key, label in items:
                if key in test:
                    item = test[key]
                    if isinstance(item, dict):
                        value = item.get("value", 0)
                        formula = item.get("formula", "")
                    else:
                        value = item
                        formula = ""
                    is_total = key == "impairment_loss"
                    row = self._write_data_row(ws, row, [label, value, formula], is_total=is_total)

            row += 1

        # ===== 租赁资本化 =====
        if lease:
            row = self._write_title(ws, row, "租赁资本化 (IFRS 16)")

            # 初始确认
            init = lease.get("initial_recognition", {})
            if init:
                row = self._write_header_row(ws, row, ["初始确认", "值", "公式"])
                if "lease_liability" in init:
                    ll = init["lease_liability"]
                    row = self._write_data_row(ws, row, ["租赁负债", ll.get("value", 0), ll.get("formula", "")])
                if "rou_asset" in init:
                    rou = init["rou_asset"]
                    row = self._write_data_row(ws, row, ["使用权资产", rou.get("value", 0), rou.get("formula", "")])

            row += 1

            # 年度明细
            schedule = lease.get("annual_schedule", [])
            if schedule:
                row = self._write_header_row(ws, row, ["年份", "期初负债", "利息", "本金", "期末负债", "折旧"])
                for item in schedule:
                    row = self._write_data_row(ws, row, [
                        item.get("year", ""),
                        item.get("opening_liability", 0),
                        item.get("interest_expense", 0),
                        item.get("principal_payment", 0),
                        item.get("closing_liability", 0),
                        item.get("depreciation", 0),
                    ])

            row += 1

            # 影响汇总
            impact = lease.get("impact_summary", {})
            if impact:
                row = self._write_header_row(ws, row, ["影响项目", "值", ""])
                for key, label in [("old_rent_expense", "原租金费用"),
                                   ("new_depreciation", "新折旧"),
                                   ("new_interest", "新利息"),
                                   ("ebitda_improvement", "EBITDA改善"),
                                   ("net_income_impact", "净利润影响")]:
                    if key in impact:
                        row = self._write_data_row(ws, row, [label, impact[key], ""])

    def write_three_statement_formula(self, result: Dict[str, Any],
                                        sheet_name: str = "三表模型(公式)"):
        """
        以公式模式写入三表模型

        输出的单元格包含真正的 Excel 公式，用户可以直接在 Excel 中修改参数

        布局:
            A列: 科目名称
            B列: 金额/公式
            C列: 参数值（可修改）

        Args:
            result: ThreeStatementModel.build() 的返回结果
            sheet_name: 工作表名称
        """
        from .cell_tracker import CellTracker
        from .formula_builder import FormulaBuilder

        ws = self.wb.create_sheet(title=sheet_name)
        tracker = CellTracker()
        builder = FormulaBuilder(tracker)

        # 设置列宽
        self._set_column_width(ws, 1, 18)  # 科目
        self._set_column_width(ws, 2, 20)  # 金额/公式
        self._set_column_width(ws, 3, 15)  # 参数

        row = 1

        # 参数填充样式
        param_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # ===== 驱动假设区 =====
        row = self._write_title(ws, row, "驱动假设 (可修改)")
        row = self._write_header_row(ws, row, ["参数", "值", "说明"])

        # 从 result 中提取假设参数
        assumptions = result.get("_meta", {}).get("assumptions", {})

        params = [
            ("growth_rate", "收入增长率", assumptions.get("growth_rate", 0.09), "0.0%"),
            ("gross_margin", "毛利率", assumptions.get("gross_margin", 0.253), "0.0%"),
            ("opex_ratio", "费用率", assumptions.get("opex_ratio", 0.097), "0.0%"),
            ("tax_rate", "税率", assumptions.get("tax_rate", 0.1386), "0.0%"),
            ("interest_rate", "利率", assumptions.get("interest_rate", 0.0233), "0.00%"),
            ("ar_days", "应收周转天数", assumptions.get("ar_days", 64), "0"),
            ("ap_days", "应付周转天数", assumptions.get("ap_days", 120), "0"),
            ("inv_days", "存货周转天数", assumptions.get("inv_days", 102), "0"),
        ]

        for name, label, value, fmt in params:
            ws.cell(row=row, column=1, value=label).border = self.thin_border
            cell = ws.cell(row=row, column=2, value=value)
            cell.border = self.thin_border
            cell.fill = param_fill
            cell.number_format = fmt
            ws.cell(row=row, column=3, value="← 可修改").font = Font(color="808080", italic=True)
            tracker.set(name, row, 2)  # 记录参数位置
            builder.register_param(name)  # 注册为参数（绝对引用）
            row += 1

        row += 1

        # ===== 基础数据区 =====
        row = self._write_title(ws, row, "基础数据")

        income = result.get("income_statement", {})
        base_inputs = income.get("revenue", {}).get("inputs", {})
        last_revenue = base_inputs.get("last_revenue", 0)

        ws.cell(row=row, column=1, value="上期收入").border = self.thin_border
        ws.cell(row=row, column=2, value=last_revenue).border = self.thin_border
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        tracker.set("last_revenue", row, 2)
        row += 1

        # 上期负债（用于利息计算）
        interest_inputs = income.get("interest", {}).get("inputs", {})
        opening_debt = interest_inputs.get("debt", 0)

        ws.cell(row=row, column=1, value="期初负债").border = self.thin_border
        ws.cell(row=row, column=2, value=opening_debt).border = self.thin_border
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        tracker.set("opening_debt", row, 2)
        row += 1

        row += 1

        # ===== 利润表（公式） =====
        row = self._write_title(ws, row, "利润表")
        row = self._write_header_row(ws, row, ["科目", "金额(万元)"])

        # 收入 = 上期 × (1 + 增长率)
        ws.cell(row=row, column=1, value="营业收入").border = self.thin_border
        ws.cell(row=row, column=2, value=builder.growth("last_revenue", "growth_rate")).border = self.thin_border
        tracker.set("revenue", row, 2)
        row += 1

        # 成本 = 收入 × (1 - 毛利率)
        ws.cell(row=row, column=1, value="营业成本").border = self.thin_border
        ws.cell(row=row, column=2, value=builder.margin("revenue", "gross_margin")).border = self.thin_border
        tracker.set("cost", row, 2)
        row += 1

        # 毛利 = 收入 - 成本
        ws.cell(row=row, column=1, value="毛利").border = self.thin_border
        ws.cell(row=row, column=2, value=builder.subtract("revenue", "cost")).border = self.thin_border
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = Font(bold=True)
        tracker.set("gross_profit", row, 2)
        row += 1

        # 费用 = 收入 × 费用率
        ws.cell(row=row, column=1, value="营业费用").border = self.thin_border
        ws.cell(row=row, column=2, value=builder.ratio("revenue", "opex_ratio")).border = self.thin_border
        tracker.set("opex", row, 2)
        row += 1

        # 营业利润 = 毛利 - 费用
        ws.cell(row=row, column=1, value="营业利润").border = self.thin_border
        ws.cell(row=row, column=2, value=builder.subtract("gross_profit", "opex")).border = self.thin_border
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = Font(bold=True)
        tracker.set("ebit", row, 2)
        row += 1

        # 利息 = 负债 × 利率
        ws.cell(row=row, column=1, value="利息费用").border = self.thin_border
        ws.cell(row=row, column=2, value=builder.ratio("opening_debt", "interest_rate")).border = self.thin_border
        tracker.set("interest", row, 2)
        row += 1

        # 税前利润 = 营业利润 - 利息
        ws.cell(row=row, column=1, value="税前利润").border = self.thin_border
        ws.cell(row=row, column=2, value=builder.subtract("ebit", "interest")).border = self.thin_border
        tracker.set("ebt", row, 2)
        row += 1

        # 所得税 = MAX(税前利润, 0) × 税率
        ws.cell(row=row, column=1, value="所得税").border = self.thin_border
        ws.cell(row=row, column=2, value=builder.tax("ebt", "tax_rate")).border = self.thin_border
        tracker.set("tax", row, 2)
        row += 1

        # 净利润 = 税前利润 - 所得税
        ws.cell(row=row, column=1, value="净利润").border = self.thin_border
        ws.cell(row=row, column=2, value=builder.subtract("ebt", "tax")).border = self.thin_border
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = Font(bold=True)
        tracker.set("net_income", row, 2)
        row += 1

        row += 1

        # ===== 营运资本（公式） =====
        row = self._write_title(ws, row, "营运资本")
        row = self._write_header_row(ws, row, ["科目", "金额(万元)"])

        # 应收账款 = 收入 / 365 × 天数
        ws.cell(row=row, column=1, value="应收账款").border = self.thin_border
        ws.cell(row=row, column=2, value=builder.turnover_days("revenue", "ar_days")).border = self.thin_border
        tracker.set("receivable", row, 2)
        row += 1

        # 应付账款 = 成本 / 365 × 天数
        ws.cell(row=row, column=1, value="应付账款").border = self.thin_border
        ws.cell(row=row, column=2, value=builder.turnover_days("cost", "ap_days")).border = self.thin_border
        tracker.set("payable", row, 2)
        row += 1

        # 存货 = 成本 / 365 × 天数
        ws.cell(row=row, column=1, value="存货").border = self.thin_border
        ws.cell(row=row, column=2, value=builder.turnover_days("cost", "inv_days")).border = self.thin_border
        tracker.set("inventory", row, 2)
        row += 1

        row += 1

        # ===== 说明 =====
        row = self._write_title(ws, row, "使用说明")
        ws.cell(row=row, column=1, value="1. 修改黄色单元格中的参数值")
        row += 1
        ws.cell(row=row, column=1, value="2. 所有公式会自动重新计算")
        row += 1
        ws.cell(row=row, column=1, value="3. 点击单元格可在公式栏查看公式")
        row += 1

    def write_lbo(self, result: Dict[str, Any], sheet_name: str = "LBO模型"):
        """
        写入 LBO 模型结果

        Args:
            result: LBOModel.build() 的返回结果
            sheet_name: 工作表名称
        """
        ws = self.wb.create_sheet(title=sheet_name)

        # 设置列宽
        self._set_column_width(ws, 1, 22)
        for i in range(2, 10):
            self._set_column_width(ws, i, 15)

        row = 1

        # ===== 元信息 =====
        meta = result.get("_meta", {})
        row = self._write_title(ws, row, f"LBO 模型 - 持有期 {meta.get('holding_period', 5)} 年")
        row += 1

        # ===== 交易结构 =====
        row = self._write_title(ws, row, "交易结构")
        row = self._write_header_row(ws, row, ["项目", "金额", "说明"])

        transaction = result.get("transaction", {})

        # 收购价格
        pp = transaction.get("purchase_price", {})
        row = self._write_data_row(ws, row, [
            "收购价格",
            pp.get("value", 0),
            pp.get("formula", "")
        ])

        # 资金来源与用途
        su = transaction.get("sources_uses", {})
        uses = su.get("uses", {})
        sources = su.get("sources", {})

        row = self._write_data_row(ws, row, ["交易费用", uses.get("transaction_fees", 0), ""])
        row = self._write_data_row(ws, row, ["融资费用", uses.get("financing_fees", 0), ""])
        row = self._write_data_row(ws, row, ["总用途", uses.get("total_uses", 0), ""], is_total=True)

        row += 1
        ws.cell(row=row, column=1, value="资金来源").font = Font(bold=True)
        row += 1

        row = self._write_data_row(ws, row, ["总债务", sources.get("total_debt", 0), ""])
        row = self._write_data_row(ws, row, ["股权投入", transaction.get("equity_invested", 0), ""], is_total=True)

        row += 1

        # ===== 运营预测 =====
        row = self._write_title(ws, row, "运营预测")

        operating_model = result.get("operating_model", {})
        operations = operating_model.get("projections", [])
        if operations:
            headers = ["指标"] + [f"Year {op['year']}" for op in operations]
            row = self._write_header_row(ws, row, headers)

            metrics = [
                ("revenue", "收入"),
                ("ebitda", "EBITDA"),
                ("ebit", "EBIT"),
                ("ufcf", "无杠杆FCF"),
            ]

            for key, label in metrics:
                values = [label] + [op.get(key, 0) for op in operations]
                is_total = key in ["ebitda", "ufcf"]
                row = self._write_data_row(ws, row, values, is_total=is_total)

        row += 1

        # ===== 债务计划 =====
        row = self._write_title(ws, row, "债务计划")

        debt_schedule = result.get("debt_schedule", {})
        annual_summary = debt_schedule.get("annual_summary", [])

        if annual_summary:
            headers = ["指标"] + [f"Year {s['year']}" for s in annual_summary]
            row = self._write_header_row(ws, row, headers)

            debt_metrics = [
                ("total_interest", "利息支出"),
                ("mandatory_amort", "强制还本"),
                ("cash_sweep", "现金扫荡"),
                ("total_debt_paydown", "总还本"),
                ("ending_debt", "期末债务"),
            ]

            for key, label in debt_metrics:
                values = [label] + [s.get(key, 0) for s in annual_summary]
                is_total = key == "ending_debt"
                row = self._write_data_row(ws, row, values, is_total=is_total)

        row += 1

        # ===== 信用指标 =====
        row = self._write_title(ws, row, "信用指标")

        credit_stats = result.get("credit_stats", [])
        if credit_stats:
            headers = ["指标"] + [f"Year {s['year']}" for s in credit_stats]
            row = self._write_header_row(ws, row, headers)

            row = self._write_data_row(ws, row,
                ["Debt/EBITDA"] + [f"{s.get('debt_to_ebitda', 0):.1f}x" for s in credit_stats])
            row = self._write_data_row(ws, row,
                ["利息覆盖倍数"] + [f"{s.get('interest_coverage', 0):.1f}x" for s in credit_stats])

        row += 1

        # ===== 退出分析 =====
        row = self._write_title(ws, row, "退出分析")
        row = self._write_header_row(ws, row, ["项目", "金额", "公式"])

        exit_analysis = result.get("exit_analysis", {})

        row = self._write_data_row(ws, row, [
            "退出EBITDA",
            exit_analysis.get("exit_ebitda", 0),
            ""
        ])

        ev = exit_analysis.get("exit_value", {})
        row = self._write_data_row(ws, row, [
            "退出价值",
            ev.get("value", 0),
            ev.get("formula", "")
        ])

        row = self._write_data_row(ws, row, [
            "期末债务",
            exit_analysis.get("ending_debt", 0),
            ""
        ])

        ep = exit_analysis.get("equity_proceeds", {})
        row = self._write_data_row(ws, row, [
            "股权所得",
            ep.get("value", 0),
            ep.get("formula", "")
        ], is_total=True)

        row += 1

        # ===== 回报指标 =====
        row = self._write_title(ws, row, "回报指标")
        row = self._write_header_row(ws, row, ["指标", "值", "公式"])

        returns = result.get("returns", {})

        irr = returns.get("irr", {})
        row = self._write_data_row(ws, row, [
            "IRR",
            f"{irr.get('value', 0):.1%}",
            irr.get("formula", "")
        ], is_total=True)

        moic = returns.get("moic", {})
        row = self._write_data_row(ws, row, [
            "MOIC",
            f"{moic.get('value', 0):.2f}x",
            moic.get("formula", "")
        ], is_total=True)

        # 现金流
        row += 1
        row = self._write_title(ws, row, "现金流序列")
        cash_flows = returns.get("cash_flows", [])
        if cash_flows:
            cf_headers = ["Year"] + [str(i) for i in range(len(cash_flows))]
            row = self._write_header_row(ws, row, cf_headers)
            cf_values = ["Cash Flow"] + cash_flows
            row = self._write_data_row(ws, row, cf_values)

    def write_lbo_sensitivity(self, result: Dict[str, Any], sheet_name: str = "LBO敏感性"):
        """
        写入 LBO 敏感性分析结果

        Args:
            result: LBOModel.sensitivity_entry_exit() 的返回结果
            sheet_name: 工作表名称
        """
        ws = self.wb.create_sheet(title=sheet_name)

        self._set_column_width(ws, 1, 18)
        for i in range(2, 8):
            self._set_column_width(ws, i, 12)

        row = 1

        # IRR 敏感性
        irr_sens = result.get("irr_sensitivity", {})
        if irr_sens:
            row = self._write_title(ws, row, "IRR 敏感性 (Entry vs Exit Multiple)")
            row += 1

            data = irr_sens.get("data", [])
            if data:
                # 表头
                first_row = data[0]
                headers = ["Entry \\ Exit"] + [k for k in first_row.keys() if k != "entry_multiple"]
                row = self._write_header_row(ws, row, headers)

                # 数据
                for d in data:
                    values = [d["entry_multiple"]] + [d.get(k, "") for k in headers[1:]]
                    row = self._write_data_row(ws, row, values)

        row += 2

        # MOIC 敏感性
        moic_sens = result.get("moic_sensitivity", {})
        if moic_sens:
            row = self._write_title(ws, row, "MOIC 敏感性 (Entry vs Exit Multiple)")
            row += 1

            data = moic_sens.get("data", [])
            if data:
                first_row = data[0]
                headers = ["Entry \\ Exit"] + [k for k in first_row.keys() if k != "entry_multiple"]
                row = self._write_header_row(ws, row, headers)

                for d in data:
                    values = [d["entry_multiple"]] + [d.get(k, "") for k in headers[1:]]
                    row = self._write_data_row(ws, row, values)

    def save(self, filepath: str):
        """
        保存 Excel 文件

        Args:
            filepath: 文件路径
        """
        self.wb.save(filepath)
        print(f"✓ Excel 已保存: {filepath}")
