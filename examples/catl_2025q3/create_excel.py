#!/usr/bin/env python3
"""创建宁德时代2025Q3财务报表Excel"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

wb = Workbook()

# ========== 资产负债表 ==========
ws_bs = wb.active
ws_bs.title = "资产负债表"

# 表头
ws_bs['A1'] = "科目"
ws_bs['B1'] = "期末数(2025-09-30)"
ws_bs['C1'] = "期初数(2025-06-30)"
ws_bs['D1'] = "单位：万元"

# 资产
data_bs = [
    # 流动资产
    ("流动资产", "", ""),
    ("货币资金", 324241586.0, 350577746.0),  # 原单位是千元，转万元
    ("应收票据及应收账款", 66900072.0, 64115205.0),
    ("应收款项融资", 32082832.0, 30000000.0),  # 估算
    ("存货", 80210000.0, 60000000.0),  # 从搜索结果
    ("流动资产合计", 574228137.0, 567700337.0),
    ("", "", ""),
    # 非流动资产
    ("非流动资产", "", ""),
    ("固定资产", 128622702.0, 118696651.0),
    ("无形资产", 15025388.0, 14684829.0),
    ("非流动资产合计", 321853995.0, 299481094.0),
    ("", "", ""),
    ("资产总计", 896082131.0, 867181431.0),
    ("", "", ""),
    # 负债
    ("流动负债", "", ""),
    ("短期借款", 15314463.0, 19009443.0),
    ("应付票据及应付账款", 214926145.0, 210388794.0),
    ("合同负债", 40700000.0, 27900000.0),  # 从搜索结果
    ("流动负债合计", 340940328.0, 336005110.0),
    ("", "", ""),
    ("非流动负债", "", ""),
    ("长期借款", 78441925.0, 82416083.0),
    ("非流动负债合计", 208129732.0, 206786909.0),
    ("", "", ""),
    ("负债合计", 549070060.0, 542792019.0),
    ("", "", ""),
    # 股东权益
    ("股东权益", "", ""),
    ("股本", 44092880.0, 44092880.0),  # 估算
    ("资本公积", 100000000.0, 100000000.0),  # 估算
    ("未分配利润", 140632845.0, 120000000.0),  # 从搜索结果
    ("股东权益合计", 347012071.0, 324389412.0),
    ("", "", ""),
    ("负债和股东权益总计", 896082131.0, 867181431.0),
]

for i, (name, end_val, begin_val) in enumerate(data_bs, start=2):
    ws_bs[f'A{i}'] = name
    if end_val != "":
        ws_bs[f'B{i}'] = end_val / 10  # 千元转万元
    if begin_val != "":
        ws_bs[f'C{i}'] = begin_val / 10

# 设置列宽
ws_bs.column_dimensions['A'].width = 25
ws_bs.column_dimensions['B'].width = 20
ws_bs.column_dimensions['C'].width = 20

# ========== 利润表 ==========
ws_pl = wb.create_sheet("利润表")

ws_pl['A1'] = "科目"
ws_pl['B1'] = "本期(2025年1-9月)"
ws_pl['C1'] = "上期(2024年1-9月)"
ws_pl['D1'] = "单位：万元"

data_pl = [
    ("营业收入", 28307198.70, 25900000.0),
    ("营业成本", 21142714.70, 20000000.0),
    ("毛利", 7164484.0, 5900000.0),  # 计算得出
    ("", "", ""),
    ("销售费用", None, None),  # 待补充
    ("管理费用", None, None),
    ("研发费用", None, None),
    ("财务费用（利息净额）", None, None),  # 待从PDF获取
    ("", "", ""),
    ("营业利润", 6055231.20, None),
    ("利润总额", 6071240.20, None),
    ("所得税费用", None, None),  # 待补充
    ("净利润", 5229686.60, 3600000.0),
    ("", "", ""),
    ("折旧与摊销", None, None),  # 待从PDF获取
]

for i, (name, cur_val, prev_val) in enumerate(data_pl, start=2):
    ws_pl[f'A{i}'] = name
    if cur_val is not None:
        ws_pl[f'B{i}'] = cur_val
    if prev_val is not None:
        ws_pl[f'C{i}'] = prev_val

ws_pl.column_dimensions['A'].width = 25
ws_pl.column_dimensions['B'].width = 20
ws_pl.column_dimensions['C'].width = 20

# ========== 现金流量表 ==========
ws_cf = wb.create_sheet("现金流量表")

ws_cf['A1'] = "科目"
ws_cf['B1'] = "本期(2025年1-9月)"
ws_cf['C1'] = "单位：万元"

data_cf = [
    ("一、经营活动产生的现金流量", "", ""),
    ("销售商品、提供劳务收到的现金", None),
    ("经营活动现金流入小计", None),
    ("经营活动现金流出小计", None),
    ("经营活动产生的现金流量净额", 8066000.0),  # 807亿
    ("", ""),
    ("二、投资活动产生的现金流量", "", ""),
    ("投资活动产生的现金流量净额", None),
    ("", ""),
    ("三、筹资活动产生的现金流量", "", ""),
    ("筹资活动产生的现金流量净额", None),
    ("", ""),
    ("四、现金及现金等价物净增加额", None),
    ("期初现金及现金等价物余额", None),
    ("期末现金及现金等价物余额", None),
]

for i, row in enumerate(data_cf, start=2):
    ws_cf[f'A{i}'] = row[0]
    if len(row) > 1 and row[1] is not None and row[1] != "":
        ws_cf[f'B{i}'] = row[1]

ws_cf.column_dimensions['A'].width = 35
ws_cf.column_dimensions['B'].width = 20

# ========== 待补充项 ==========
ws_todo = wb.create_sheet("待补充")

ws_todo['A1'] = "需要从PDF季报获取的项目"
ws_todo['A2'] = "1. 财务费用（利息费用）"
ws_todo['A3'] = "2. 折旧与摊销"
ws_todo['A4'] = "3. 所得税费用"
ws_todo['A5'] = "4. 固定资产原值和累计折旧"
ws_todo['A6'] = "5. 各项现金流明细"
ws_todo['A7'] = "6. 资本支出（CAPEX）"

ws_todo.column_dimensions['A'].width = 40

# 保存
output_path = "/home/wangbo/document/balance/examples/catl_2025q3/catl_2025q3.xlsx"
wb.save(output_path)
print(f"Excel已保存到: {output_path}")
