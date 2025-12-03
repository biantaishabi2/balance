#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fm - Financial Model CLI

统一的财务建模命令行工具，整合 LBO、M&A、DCF、三表模型等功能。

用法:
    fm <command> [subcommand] [options] < input.json

命令:
    lbo      杠杆收购分析
    ma       并购分析
    dcf      DCF估值
    three    三表模型
    tool     原子工具
"""

import sys
import json
import argparse
from typing import Dict, Any, List, Optional


# ============================================================
# 输出格式化
# ============================================================

def format_number(value: float, style: str = "auto") -> str:
    """格式化数字"""
    if value is None:
        return "N/A"

    abs_val = abs(value)

    if style == "percent":
        return f"{value:.2%}"
    elif style == "multiple":
        return f"{value:.2f}x"
    elif style == "currency" or (style == "auto" and abs_val >= 1000):
        if abs_val >= 1e9:
            return f"{value/1e9:,.2f}B"
        elif abs_val >= 1e6:
            return f"{value/1e6:,.2f}M"
        elif abs_val >= 1e3:
            return f"{value/1e3:,.2f}K"
        else:
            return f"{value:,.2f}"
    else:
        return f"{value:.4f}"


def print_json(data: Dict, compact: bool = False):
    """输出JSON"""
    if compact:
        print(json.dumps(data, ensure_ascii=False))
    else:
        print(json.dumps(data, indent=2, ensure_ascii=False))


def print_table(headers: List[str], rows: List[List[str]], title: str = None):
    """输出表格"""
    if title:
        print(f"\n{title}")
        print("─" * 60)

    # 计算列宽
    widths = [len(h) for h in headers]
    for row in rows:
        for i, cell in enumerate(row):
            widths[i] = max(widths[i], len(str(cell)))

    # 打印表头
    header_line = "│ " + " │ ".join(h.ljust(widths[i]) for i, h in enumerate(headers)) + " │"
    separator = "├─" + "─┼─".join("─" * w for w in widths) + "─┤"
    top_border = "┌─" + "─┬─".join("─" * w for w in widths) + "─┐"
    bottom_border = "└─" + "─┴─".join("─" * w for w in widths) + "─┘"

    print(top_border)
    print(header_line)
    print(separator)

    # 打印数据行
    for row in rows:
        row_line = "│ " + " │ ".join(str(cell).ljust(widths[i]) for i, cell in enumerate(row)) + " │"
        print(row_line)

    print(bottom_border)


def read_json_input() -> Dict:
    """从stdin读取JSON"""
    try:
        return json.load(sys.stdin)
    except json.JSONDecodeError as e:
        print(f"错误: 无效的JSON输入 - {e}", file=sys.stderr)
        sys.exit(1)


def parse_float_list(value: str) -> List[float]:
    """解析逗号分隔的浮点数列表"""
    return [float(x.strip()) for x in value.split(",")]


# ============================================================
# LBO 命令
# ============================================================

def cmd_lbo_calc(args):
    """LBO计算"""
    from financial_model.tools import lbo_quick_build

    data = read_json_input()

    # 覆盖参数
    if args.entry_multiple:
        data["entry_multiple"] = args.entry_multiple
    if args.exit_multiple:
        data["exit_multiple"] = args.exit_multiple
    if args.holding_period:
        data["holding_period"] = args.holding_period
    if args.sweep_percent:
        data["sweep_percent"] = args.sweep_percent

    result = lbo_quick_build(data)

    if args.detail:
        print_json(result, args.compact)
    else:
        # 简化输出
        summary = {
            "irr": result["returns"]["irr"]["value"],
            "irr_formatted": format_number(result["returns"]["irr"]["value"], "percent"),
            "moic": result["returns"]["moic"]["value"],
            "moic_formatted": format_number(result["returns"]["moic"]["value"], "multiple"),
            "equity_invested": result["transaction"]["equity_invested"],
            "equity_proceeds": result["exit_analysis"]["equity_proceeds"]["value"],
            "exit_debt": result["exit_analysis"]["ending_debt"],
            "holding_period": data.get("holding_period", 5)
        }
        print_json(summary, args.compact)


def cmd_lbo_sensitivity(args):
    """LBO敏感性分析"""
    from financial_model.tools import lbo_quick_build

    data = read_json_input()

    entry_range = parse_float_list(args.entry) if args.entry else [7.0, 7.5, 8.0, 8.5, 9.0]
    exit_range = parse_float_list(args.exit) if args.exit else [7.0, 8.0, 9.0]

    metric = args.metric or "irr"

    # 构建矩阵
    rows = []
    for entry in entry_range:
        row = [f"{entry:.1f}x"]
        for exit_mult in exit_range:
            inputs = data.copy()
            inputs["entry_multiple"] = entry
            inputs["exit_multiple"] = exit_mult

            result = lbo_quick_build(inputs)

            if metric == "irr":
                value = result["returns"]["irr"]["value"]
                row.append(format_number(value, "percent"))
            else:
                value = result["returns"]["moic"]["value"]
                row.append(format_number(value, "multiple"))
        rows.append(row)

    headers = ["Entry \\ Exit"] + [f"{x:.1f}x" for x in exit_range]

    if args.format == "json":
        output = {
            "metric": metric.upper(),
            "entry_range": entry_range,
            "exit_range": exit_range,
            "matrix": rows
        }
        print_json(output, args.compact)
    else:
        print_table(headers, rows, f"{metric.upper()} Sensitivity (Entry vs Exit Multiple)")


def cmd_lbo_export(args):
    """LBO导出Excel"""
    from financial_model.models import LBOModel

    data = read_json_input()

    lbo = LBOModel()

    # 如果输入是完整结果，直接导出；否则先计算
    if "_meta" in data and data["_meta"].get("model_type") == "lbo":
        result = data
    else:
        result = lbo.build(data)

    # 导出（需要实现to_excel方法）
    output_file = args.output or "lbo_model.xlsx"

    try:
        # 简单导出逻辑
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LBO Summary"

        ws["A1"] = "LBO Analysis"
        ws["A3"] = "IRR"
        ws["B3"] = result["returns"]["irr"]["value"]
        ws["A4"] = "MOIC"
        ws["B4"] = result["returns"]["moic"]["value"]
        ws["A5"] = "Equity Invested"
        ws["B5"] = result["transaction"]["equity_invested"]
        ws["A6"] = "Equity Proceeds"
        ws["B6"] = result["exit_analysis"]["equity_proceeds"]["value"]

        wb.save(output_file)
        print(f"已导出到 {output_file}", file=sys.stderr)
    except ImportError:
        print("错误: 需要安装 openpyxl (pip install openpyxl)", file=sys.stderr)
        sys.exit(1)


# ============================================================
# M&A 命令
# ============================================================

def cmd_ma_calc(args):
    """M&A计算"""
    from financial_model.tools import ma_quick_build

    data = read_json_input()

    # 覆盖参数
    if args.premium:
        data.setdefault("deal_terms", {})["premium_percent"] = args.premium
    if args.cash_percent:
        data.setdefault("deal_terms", {})["cash_percent"] = args.cash_percent
    if args.stock_percent:
        data.setdefault("deal_terms", {})["stock_percent"] = args.stock_percent
    if args.synergies:
        data.setdefault("synergies", {})["cost_synergies"] = args.synergies

    result = ma_quick_build(data)

    if args.detail:
        print_json(result, args.compact)
    else:
        ad = result["accretion_dilution"]["without_synergies"]["accretion_dilution"]
        ad_with = None
        if result["accretion_dilution"].get("with_synergies"):
            ad_with = result["accretion_dilution"]["with_synergies"]["accretion_dilution"]

        summary = {
            "purchase_price": result["deal_summary"]["purchase_price"],
            "premium": format_number(result["deal_summary"]["premium"]["percent"], "percent"),
            "goodwill": result["purchase_price_allocation"]["goodwill"],
            "eps_impact": {
                "without_synergies": format_number(ad["percent"], "percent"),
                "status": ad["status"]
            },
            "breakeven_synergies": result["breakeven"]["synergies_needed"]
        }

        if ad_with:
            summary["eps_impact"]["with_synergies"] = format_number(ad_with["percent"], "percent")

        print_json(summary, args.compact)


def cmd_ma_accretion(args):
    """M&A增厚/稀释敏感性"""
    from financial_model.tools import ma_quick_build

    data = read_json_input()

    premium_range = parse_float_list(args.premium) if args.premium else [0.20, 0.30, 0.40]
    stock_range = parse_float_list(args.stock) if args.stock else [0.20, 0.40, 0.60]

    rows = []
    for premium in premium_range:
        row = [format_number(premium, "percent")]
        for stock_pct in stock_range:
            inputs = data.copy()
            inputs.setdefault("deal_terms", {})
            inputs["deal_terms"]["premium_percent"] = premium
            inputs["deal_terms"]["stock_percent"] = stock_pct
            inputs["deal_terms"]["cash_percent"] = 1.0 - stock_pct

            result = ma_quick_build(inputs)
            ad = result["accretion_dilution"]["without_synergies"]["accretion_dilution"]
            row.append(format_number(ad["percent"], "percent"))
        rows.append(row)

    headers = ["Premium \\ Stock"] + [format_number(x, "percent") for x in stock_range]

    if args.format == "json":
        output = {
            "metric": "EPS Accretion/Dilution",
            "premium_range": premium_range,
            "stock_range": stock_range,
            "matrix": rows
        }
        print_json(output, args.compact)
    else:
        print_table(headers, rows, "EPS Accretion/Dilution (Premium vs Stock %)")


def cmd_ma_breakeven(args):
    """M&A盈亏平衡"""
    from financial_model.tools import ma_quick_build

    data = read_json_input()
    result = ma_quick_build(data)

    be = result["breakeven"]
    output = {
        "synergies_needed": be["synergies_needed"],
        "synergies_formatted": format_number(be["synergies_needed"], "currency"),
        "current_status": be["current_dilution"]["status"],
        "current_eps_impact": format_number(be["current_dilution"]["percent"], "percent"),
        "interpretation": be["interpretation"]
    }

    print_json(output, args.compact)


def cmd_ma_export(args):
    """M&A导出Excel"""
    from financial_model.models import MAModel

    data = read_json_input()

    ma = MAModel()

    if "_meta" in data and data["_meta"].get("model_type") == "ma":
        result = data
    else:
        result = ma.build(data)

    output_file = args.output or "ma_model.xlsx"
    ma.to_excel(result, output_file)
    print(f"已导出到 {output_file}", file=sys.stderr)


# ============================================================
# DCF 命令
# ============================================================

def cmd_dcf_calc(args):
    """DCF计算"""
    from financial_model.tools import (
        calc_wacc, calc_terminal_value, calc_enterprise_value,
        calc_equity_value, calc_per_share_value, dcf_quick_valuate
    )

    data = read_json_input()

    # 覆盖参数
    if args.wacc:
        data["wacc"] = args.wacc
    if args.growth:
        data["terminal_growth"] = args.growth

    # 简单模式：直接提供FCF列表
    if "fcf_projections" in data or "fcf_list" in data:
        fcf_list = data.get("fcf_projections") or data.get("fcf_list")
        wacc = data.get("wacc", 0.10)
        growth = data.get("terminal_growth", 0.02)
        net_debt = data.get("net_debt", 0)
        shares = data.get("shares_outstanding", 1)

        # 终值
        final_fcf = fcf_list[-1]
        tv = calc_terminal_value(
            method="perpetual",
            final_fcf=final_fcf,
            wacc=wacc,
            terminal_growth=growth
        )

        # 企业价值
        ev = calc_enterprise_value(
            fcf_list=fcf_list,
            wacc=wacc,
            terminal_value=tv["value"]
        )

        # 股权价值
        equity = calc_equity_value(
            enterprise_value=ev["value"],
            debt=net_debt,
            cash=0
        )

        # 每股价值
        per_share = calc_per_share_value(
            equity_value=equity["value"],
            shares_outstanding=shares
        )

        result = {
            "enterprise_value": ev["value"],
            "equity_value": equity["value"],
            "per_share_value": per_share["value"],
            "terminal_value": tv["value"],
            "pv_of_fcf": ev["inputs"]["pv_of_fcf"],
            "pv_of_terminal": ev["inputs"]["pv_of_terminal"],
            "wacc": wacc,
            "terminal_growth": growth
        }
    else:
        # 完整模式
        result = dcf_quick_valuate(data)

    print_json(result, args.compact)


def cmd_dcf_sensitivity(args):
    """DCF敏感性分析"""
    from financial_model.tools import dcf_sensitivity

    data = read_json_input()

    wacc_range = parse_float_list(args.wacc) if args.wacc else [0.08, 0.09, 0.10, 0.11]
    growth_range = parse_float_list(args.growth) if args.growth else [0.01, 0.02, 0.03]

    fcf_list = data.get("fcf_projections") or data.get("fcf_list")
    net_debt = data.get("net_debt", 0)
    shares = data.get("shares_outstanding", 1)

    result = dcf_sensitivity(
        fcf_list=fcf_list,
        wacc_range=wacc_range,
        growth_range=growth_range,
        debt=net_debt,
        cash=0,
        shares=shares
    )

    if args.format == "json":
        print_json(result, args.compact)
    else:
        # 表格输出
        headers = ["WACC \\ Growth"] + [format_number(g, "percent") for g in growth_range]
        rows = []
        for row_data in result["matrix"]:
            row = [row_data["wacc"]]
            for g in growth_range:
                key = f"g={g:.1%)}".replace(")", "")
                # 查找对应值
                for k, v in row_data.items():
                    if k.startswith("g="):
                        row.append(v)
                        break
            rows.append(row)

        print_table(headers, rows, "Per Share Value Sensitivity (WACC vs Growth)")


def cmd_dcf_wacc(args):
    """WACC计算器"""
    from financial_model.tools import calc_capm, calc_wacc

    # CAPM
    cost_of_equity = calc_capm(
        risk_free_rate=args.rf,
        beta=args.beta,
        market_risk_premium=args.mrp
    )

    # WACC
    wacc = calc_wacc(
        cost_of_equity=cost_of_equity["value"],
        cost_of_debt=args.rd,
        tax_rate=args.tax,
        debt_ratio=args.de
    )

    result = {
        "cost_of_equity": cost_of_equity["value"],
        "cost_of_equity_formatted": format_number(cost_of_equity["value"], "percent"),
        "cost_of_debt_after_tax": args.rd * (1 - args.tax),
        "wacc": wacc["value"],
        "wacc_formatted": format_number(wacc["value"], "percent"),
        "inputs": {
            "risk_free_rate": args.rf,
            "beta": args.beta,
            "market_risk_premium": args.mrp,
            "cost_of_debt": args.rd,
            "tax_rate": args.tax,
            "debt_ratio": args.de
        }
    }

    print_json(result, False)


# ============================================================
# Three Statement 命令
# ============================================================

def cmd_three_forecast(args):
    """三表预测"""
    from financial_model.tools import three_statement_quick_build

    data = read_json_input()

    if args.years:
        data.setdefault("assumptions", {})["years"] = args.years
    if args.growth:
        data.setdefault("assumptions", {})["revenue_growth"] = args.growth

    result = three_statement_quick_build(data)
    print_json(result, args.compact)


def cmd_three_check(args):
    """三表配平检验"""
    from financial_model.tools import check_balance

    data = read_json_input()

    # 提取资产负债表数据
    total_assets = data.get("total_assets", 0)
    total_liabilities = data.get("total_liabilities", 0)
    total_equity = data.get("total_equity", 0)

    result = check_balance(
        total_assets=total_assets,
        total_liabilities=total_liabilities,
        total_equity=total_equity
    )

    print_json(result, args.compact)


# ============================================================
# Ratio 命令
# ============================================================

def cmd_ratio_calc(args):
    """比率计算"""
    from financial_model.tools import (
        calc_profitability, calc_liquidity, calc_solvency,
        calc_efficiency, calc_valuation, calc_all_ratios
    )

    data = read_json_input()

    # 从输入数据提取财务数据
    income = data.get("income_statement", data)
    balance = data.get("balance_sheet", data)
    market = data.get("market_data", data)

    ratio_type = args.type or "all"

    if ratio_type == "all":
        result = calc_all_ratios(
            income_statement=income,
            balance_sheet=balance,
            market_data=market
        )
    elif ratio_type == "profitability":
        result = calc_profitability(
            revenue=income.get("revenue"),
            gross_profit=income.get("gross_profit"),
            cost_of_revenue=income.get("cost_of_revenue"),
            operating_income=income.get("operating_income"),
            net_income=income.get("net_income"),
            ebitda=income.get("ebitda"),
            total_assets=balance.get("total_assets"),
            total_equity=balance.get("total_equity")
        )
    elif ratio_type == "liquidity":
        result = calc_liquidity(
            cash=balance.get("cash"),
            inventory=balance.get("inventory"),
            current_assets=balance.get("current_assets"),
            current_liabilities=balance.get("current_liabilities")
        )
    elif ratio_type == "solvency":
        total_debt = balance.get("total_debt")
        if total_debt is None:
            total_debt = balance.get("short_term_debt", 0) + balance.get("long_term_debt", 0)
        result = calc_solvency(
            total_debt=total_debt,
            total_equity=balance.get("total_equity"),
            total_assets=balance.get("total_assets"),
            ebit=income.get("ebit"),
            ebitda=income.get("ebitda"),
            interest_expense=income.get("interest_expense")
        )
    elif ratio_type == "efficiency":
        result = calc_efficiency(
            revenue=income.get("revenue"),
            cost_of_revenue=income.get("cost_of_revenue"),
            total_assets=balance.get("total_assets"),
            inventory=balance.get("inventory"),
            accounts_receivable=balance.get("accounts_receivable"),
            accounts_payable=balance.get("accounts_payable")
        )
    elif ratio_type == "valuation":
        total_debt = balance.get("total_debt")
        if total_debt is None:
            total_debt = balance.get("short_term_debt", 0) + balance.get("long_term_debt", 0)
        result = calc_valuation(
            market_cap=market.get("market_cap"),
            stock_price=market.get("stock_price"),
            shares_outstanding=market.get("shares_outstanding"),
            net_income=income.get("net_income"),
            total_equity=balance.get("total_equity"),
            revenue=income.get("revenue"),
            ebitda=income.get("ebitda"),
            total_debt=total_debt,
            cash=balance.get("cash")
        )
    else:
        print(f"错误: 未知的比率类型 '{ratio_type}'", file=sys.stderr)
        sys.exit(1)

    print_json(result, args.compact)


def cmd_ratio_dupont(args):
    """杜邦分析"""
    from financial_model.tools import calc_dupont

    data = read_json_input()

    income = data.get("income_statement", data)
    balance = data.get("balance_sheet", data)

    levels = args.levels or 3

    result = calc_dupont(
        net_income=income.get("net_income"),
        revenue=income.get("revenue"),
        total_assets=balance.get("total_assets"),
        total_equity=balance.get("total_equity"),
        ebit=income.get("ebit"),
        ebt=income.get("ebt"),
        levels=levels
    )

    print_json(result, args.compact)


def cmd_ratio_compare(args):
    """同业对比"""
    data = read_json_input()

    # 输入应该是多个公司的数据
    companies = data.get("companies", [])
    metrics = args.metrics.split(",") if args.metrics else ["roe", "net_margin", "roic"]

    from financial_model.tools import calc_profitability, calc_solvency

    comparison = {"metrics": {}}

    for metric in metrics:
        comparison["metrics"][metric] = {}

        for company in companies:
            name = company.get("name", "Unknown")
            income = company.get("income_statement", company)
            balance = company.get("balance_sheet", company)

            # 计算指标
            if metric in ["gross_margin", "operating_margin", "net_margin", "ebitda_margin", "roa", "roe", "roic"]:
                result = calc_profitability(
                    revenue=income.get("revenue"),
                    gross_profit=income.get("gross_profit"),
                    operating_income=income.get("operating_income"),
                    net_income=income.get("net_income"),
                    ebitda=income.get("ebitda"),
                    total_assets=balance.get("total_assets"),
                    total_equity=balance.get("total_equity")
                )
                if metric in result.get("ratios", {}):
                    comparison["metrics"][metric][name] = result["ratios"][metric]["value"]

            elif metric in ["debt_to_equity", "interest_coverage"]:
                result = calc_solvency(
                    total_debt=balance.get("total_debt"),
                    total_equity=balance.get("total_equity"),
                    total_assets=balance.get("total_assets"),
                    ebit=income.get("ebit"),
                    interest_expense=income.get("interest_expense")
                )
                if metric in result.get("ratios", {}):
                    comparison["metrics"][metric][name] = result["ratios"][metric]["value"]

    # 生成排名
    comparison["ranking"] = {}
    for metric, values in comparison["metrics"].items():
        sorted_companies = sorted(values.items(), key=lambda x: x[1], reverse=True)
        comparison["ranking"][metric] = [c[0] for c in sorted_companies]

    print_json(comparison, args.compact)


def cmd_ratio_trend(args):
    """趋势分析"""
    from financial_model.tools import calc_profitability

    data = read_json_input()

    # 输入应该是多期数据
    periods = data.get("periods", [])
    metrics = args.metrics.split(",") if args.metrics else ["roe", "net_margin", "asset_turnover"]

    trend = {"trend_analysis": {}}

    for metric in metrics:
        values = []

        for period_data in periods:
            income = period_data.get("income_statement", period_data)
            balance = period_data.get("balance_sheet", period_data)

            result = calc_profitability(
                revenue=income.get("revenue"),
                gross_profit=income.get("gross_profit"),
                operating_income=income.get("operating_income"),
                net_income=income.get("net_income"),
                total_assets=balance.get("total_assets"),
                total_equity=balance.get("total_equity")
            )

            if metric in result.get("ratios", {}):
                values.append(result["ratios"][metric]["value"])

        if values:
            # 计算CAGR
            if len(values) > 1 and values[0] > 0:
                cagr = (values[-1] / values[0]) ** (1 / (len(values) - 1)) - 1
            else:
                cagr = 0

            # 判断趋势
            if len(values) > 1:
                if values[-1] > values[0] * 1.1:
                    trend_desc = "上升"
                elif values[-1] < values[0] * 0.9:
                    trend_desc = "下降"
                else:
                    trend_desc = "稳定"
            else:
                trend_desc = "数据不足"

            trend["trend_analysis"][metric] = {
                "values": values,
                "cagr": round(cagr, 4),
                "trend": trend_desc
            }

    print_json(trend, args.compact)


# ============================================================
# Prepare 命令
# ============================================================

def cmd_prepare_lbo(args):
    """准备LBO数据"""
    from financial_model.tools import prepare_lbo_data

    data = read_json_input()

    # 兼容不同字段名
    transaction = data.get("transaction") or data.get("transaction_assumptions", {})

    result = prepare_lbo_data(
        income_statement=data.get("income_statement", {}),
        balance_sheet=data.get("balance_sheet", {}),
        transaction=transaction,
        debt_structure=data.get("debt_structure"),
        operating_assumptions=data.get("operating_assumptions")
    )

    print_json(result, args.compact)


def cmd_prepare_ma(args):
    """准备MA数据"""
    from financial_model.tools import prepare_ma_data

    data = read_json_input()

    result = prepare_ma_data(
        acquirer=data.get("acquirer", {}),
        target=data.get("target", {}),
        deal_terms=data.get("deal_terms"),
        synergies=data.get("synergies")
    )

    print_json(result, args.compact)


def cmd_prepare_dcf(args):
    """准备DCF数据"""
    from financial_model.tools import prepare_dcf_data

    data = read_json_input()

    result = prepare_dcf_data(
        income_statement=data.get("income_statement", {}),
        balance_sheet=data.get("balance_sheet", {}),
        projections=data.get("projections"),
        valuation=data.get("valuation")
    )

    print_json(result, args.compact)


# ============================================================
# Tool 命令
# ============================================================

TOOL_REGISTRY = {
    # LBO
    "calc_purchase_price": ("financial_model.tools.lbo_tools", "计算收购价格"),
    "calc_sources_uses": ("financial_model.tools.lbo_tools", "资金来源与用途"),
    "project_operations": ("financial_model.tools.lbo_tools", "预测运营数据"),
    "calc_debt_schedule": ("financial_model.tools.lbo_tools", "债务计划表"),
    "calc_exit_value": ("financial_model.tools.lbo_tools", "退出价值"),
    "calc_equity_proceeds": ("financial_model.tools.lbo_tools", "股权所得"),
    "calc_irr": ("financial_model.tools.lbo_tools", "内部收益率"),
    "calc_moic": ("financial_model.tools.lbo_tools", "投资倍数"),
    "lbo_quick_build": ("financial_model.tools.lbo_tools", "LBO快捷构建"),
    # M&A
    "calc_offer_price": ("financial_model.tools.ma_tools", "收购报价"),
    "calc_funding_mix": ("financial_model.tools.ma_tools", "融资结构"),
    "calc_goodwill": ("financial_model.tools.ma_tools", "商誉计算"),
    "calc_pro_forma": ("financial_model.tools.ma_tools", "合并报表"),
    "calc_accretion_dilution": ("financial_model.tools.ma_tools", "增厚/稀释"),
    "calc_synergies": ("financial_model.tools.ma_tools", "协同效应"),
    "calc_breakeven": ("financial_model.tools.ma_tools", "盈亏平衡"),
    "ma_quick_build": ("financial_model.tools.ma_tools", "M&A快捷构建"),
    # DCF
    "calc_capm": ("financial_model.tools.dcf_tools", "CAPM股权成本"),
    "calc_wacc": ("financial_model.tools.dcf_tools", "加权平均资本成本"),
    "calc_fcff": ("financial_model.tools.dcf_tools", "企业自由现金流"),
    "calc_terminal_value": ("financial_model.tools.dcf_tools", "终值计算"),
    "calc_enterprise_value": ("financial_model.tools.dcf_tools", "企业价值"),
    "calc_equity_value": ("financial_model.tools.dcf_tools", "股权价值"),
    "calc_per_share_value": ("financial_model.tools.dcf_tools", "每股价值"),
    "dcf_sensitivity": ("financial_model.tools.dcf_tools", "DCF敏感性分析"),
    "dcf_quick_valuate": ("financial_model.tools.dcf_tools", "DCF快捷估值"),
    # Three Statement
    "forecast_revenue": ("financial_model.tools.three_statement_tools", "收入预测"),
    "forecast_cost": ("financial_model.tools.three_statement_tools", "成本预测"),
    "forecast_opex": ("financial_model.tools.three_statement_tools", "费用预测"),
    "calc_income_statement": ("financial_model.tools.three_statement_tools", "利润表计算"),
    "calc_working_capital": ("financial_model.tools.three_statement_tools", "营运资本"),
    "calc_depreciation": ("financial_model.tools.three_statement_tools", "折旧计算"),
    "calc_capex": ("financial_model.tools.three_statement_tools", "资本支出"),
    "calc_operating_cash_flow": ("financial_model.tools.three_statement_tools", "经营现金流"),
    "calc_cash_flow_statement": ("financial_model.tools.three_statement_tools", "现金流量表"),
    "check_balance": ("financial_model.tools.three_statement_tools", "配平检验"),
    "three_statement_quick_build": ("financial_model.tools.three_statement_tools", "三表快捷构建"),
    # Ratio Analysis
    "calc_profitability": ("financial_model.tools.ratio_tools", "盈利能力分析"),
    "calc_liquidity": ("financial_model.tools.ratio_tools", "流动性分析"),
    "calc_solvency": ("financial_model.tools.ratio_tools", "偿债能力分析"),
    "calc_efficiency": ("financial_model.tools.ratio_tools", "运营效率分析"),
    "calc_valuation": ("financial_model.tools.ratio_tools", "估值分析"),
    "calc_dupont": ("financial_model.tools.ratio_tools", "杜邦分析"),
    "calc_all_ratios": ("financial_model.tools.ratio_tools", "全部比率分析"),
    # Prepare (数据准备)
    "prepare_lbo_data": ("financial_model.tools.prepare_tools", "LBO数据准备"),
    "prepare_ma_data": ("financial_model.tools.prepare_tools", "MA数据准备"),
    "prepare_dcf_data": ("financial_model.tools.prepare_tools", "DCF数据准备"),
}


def cmd_tool_list(args):
    """列出所有工具"""
    # 按模块分组
    modules = {}
    for name, (module, desc) in TOOL_REGISTRY.items():
        mod_name = module.split(".")[-1].replace("_tools", "").upper()
        if mod_name not in modules:
            modules[mod_name] = []
        modules[mod_name].append((name, desc))

    if args.format == "json":
        print_json({"tools": TOOL_REGISTRY}, args.compact)
    else:
        print(f"\n可用工具 (共 {len(TOOL_REGISTRY)} 个)\n")
        for mod_name, tools in modules.items():
            print(f"{mod_name} Tools ({len(tools)}):")
            for name, desc in tools:
                print(f"  {name:30} {desc}")
            print()


def cmd_tool_run(args):
    """运行指定工具"""
    tool_name = args.tool_name

    if tool_name not in TOOL_REGISTRY:
        print(f"错误: 未知工具 '{tool_name}'", file=sys.stderr)
        print(f"使用 'fm tool list' 查看可用工具", file=sys.stderr)
        sys.exit(1)

    module_path, _ = TOOL_REGISTRY[tool_name]

    # 动态导入
    import importlib
    module = importlib.import_module(module_path)
    func = getattr(module, tool_name)

    # 读取输入
    data = read_json_input()

    # 调用工具
    result = func(**data)

    print_json(result, args.compact)


def cmd_tool_help(args):
    """显示工具帮助"""
    tool_name = args.tool_name

    if tool_name not in TOOL_REGISTRY:
        print(f"错误: 未知工具 '{tool_name}'", file=sys.stderr)
        sys.exit(1)

    module_path, desc = TOOL_REGISTRY[tool_name]

    # 动态导入获取docstring
    import importlib
    module = importlib.import_module(module_path)
    func = getattr(module, tool_name)

    print(f"\n{tool_name} - {desc}\n")
    if func.__doc__:
        print(func.__doc__)
    else:
        print("(无详细文档)")


# ============================================================
# 主函数
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        prog="fm",
        description="Financial Model CLI - 财务建模命令行工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  fm lbo calc < input.json
  fm lbo sensitivity --entry 7,8,9 --exit 7,8,9 < input.json
  fm ma calc < deal.json
  fm ma breakeven < deal.json
  fm dcf calc --wacc 0.09 < projections.json
  fm ratio calc --type all < financial.json
  fm ratio dupont --levels 5 < financial.json
  fm prepare lbo < financial.json | fm lbo calc
  fm prepare dcf < financial.json | fm dcf calc
  fm tool list
  fm tool run calc_irr < cashflows.json

更多帮助:
  fm <command> --help
        """
    )

    subparsers = parser.add_subparsers(dest="command", help="子命令")

    # ===== LBO =====
    lbo_parser = subparsers.add_parser("lbo", help="杠杆收购分析")
    lbo_sub = lbo_parser.add_subparsers(dest="subcommand")

    # lbo calc
    lbo_calc = lbo_sub.add_parser("calc", help="LBO计算")
    lbo_calc.add_argument("--entry-multiple", type=float, help="入场倍数")
    lbo_calc.add_argument("--exit-multiple", type=float, help="退出倍数")
    lbo_calc.add_argument("--holding-period", type=int, help="持有期(年)")
    lbo_calc.add_argument("--sweep-percent", type=float, help="现金扫荡比例")
    lbo_calc.add_argument("--detail", action="store_true", help="输出完整详情")
    lbo_calc.add_argument("--compact", action="store_true", help="紧凑输出")
    lbo_calc.set_defaults(func=cmd_lbo_calc)

    # lbo sensitivity
    lbo_sens = lbo_sub.add_parser("sensitivity", help="敏感性分析")
    lbo_sens.add_argument("--entry", help="入场倍数范围(逗号分隔)")
    lbo_sens.add_argument("--exit", help="退出倍数范围(逗号分隔)")
    lbo_sens.add_argument("--metric", choices=["irr", "moic"], default="irr", help="输出指标")
    lbo_sens.add_argument("--format", choices=["json", "table"], default="table", help="输出格式")
    lbo_sens.add_argument("--compact", action="store_true", help="紧凑输出")
    lbo_sens.set_defaults(func=cmd_lbo_sensitivity)

    # lbo export
    lbo_export = lbo_sub.add_parser("export", help="导出Excel")
    lbo_export.add_argument("-o", "--output", help="输出文件")
    lbo_export.set_defaults(func=cmd_lbo_export)

    # ===== M&A =====
    ma_parser = subparsers.add_parser("ma", help="并购分析")
    ma_sub = ma_parser.add_subparsers(dest="subcommand")

    # ma calc
    ma_calc = ma_sub.add_parser("calc", help="M&A计算")
    ma_calc.add_argument("--premium", type=float, help="溢价比例")
    ma_calc.add_argument("--cash-percent", type=float, help="现金支付比例")
    ma_calc.add_argument("--stock-percent", type=float, help="股票支付比例")
    ma_calc.add_argument("--synergies", type=float, help="协同效应金额")
    ma_calc.add_argument("--detail", action="store_true", help="输出完整详情")
    ma_calc.add_argument("--compact", action="store_true", help="紧凑输出")
    ma_calc.set_defaults(func=cmd_ma_calc)

    # ma accretion
    ma_accr = ma_sub.add_parser("accretion", help="增厚/稀释敏感性")
    ma_accr.add_argument("--premium", help="溢价范围(逗号分隔)")
    ma_accr.add_argument("--stock", help="股票比例范围(逗号分隔)")
    ma_accr.add_argument("--format", choices=["json", "table"], default="table", help="输出格式")
    ma_accr.add_argument("--compact", action="store_true", help="紧凑输出")
    ma_accr.set_defaults(func=cmd_ma_accretion)

    # ma breakeven
    ma_be = ma_sub.add_parser("breakeven", help="盈亏平衡分析")
    ma_be.add_argument("--compact", action="store_true", help="紧凑输出")
    ma_be.set_defaults(func=cmd_ma_breakeven)

    # ma export
    ma_export = ma_sub.add_parser("export", help="导出Excel")
    ma_export.add_argument("-o", "--output", help="输出文件")
    ma_export.set_defaults(func=cmd_ma_export)

    # ===== DCF =====
    dcf_parser = subparsers.add_parser("dcf", help="DCF估值")
    dcf_sub = dcf_parser.add_subparsers(dest="subcommand")

    # dcf calc
    dcf_calc = dcf_sub.add_parser("calc", help="DCF计算")
    dcf_calc.add_argument("--wacc", type=float, help="WACC")
    dcf_calc.add_argument("--growth", type=float, help="永续增长率")
    dcf_calc.add_argument("--compact", action="store_true", help="紧凑输出")
    dcf_calc.set_defaults(func=cmd_dcf_calc)

    # dcf sensitivity
    dcf_sens = dcf_sub.add_parser("sensitivity", help="敏感性分析")
    dcf_sens.add_argument("--wacc", help="WACC范围(逗号分隔)")
    dcf_sens.add_argument("--growth", help="增长率范围(逗号分隔)")
    dcf_sens.add_argument("--format", choices=["json", "table"], default="table", help="输出格式")
    dcf_sens.add_argument("--compact", action="store_true", help="紧凑输出")
    dcf_sens.set_defaults(func=cmd_dcf_sensitivity)

    # dcf wacc
    dcf_wacc = dcf_sub.add_parser("wacc", help="WACC计算器")
    dcf_wacc.add_argument("--rf", type=float, required=True, help="无风险利率")
    dcf_wacc.add_argument("--beta", type=float, required=True, help="Beta")
    dcf_wacc.add_argument("--mrp", type=float, required=True, help="市场风险溢价")
    dcf_wacc.add_argument("--rd", type=float, required=True, help="债务成本")
    dcf_wacc.add_argument("--tax", type=float, required=True, help="税率")
    dcf_wacc.add_argument("--de", type=float, required=True, help="债务比例 D/(D+E)")
    dcf_wacc.set_defaults(func=cmd_dcf_wacc)

    # ===== Three Statement =====
    three_parser = subparsers.add_parser("three", help="三表模型")
    three_sub = three_parser.add_subparsers(dest="subcommand")

    # three forecast
    three_fc = three_sub.add_parser("forecast", help="财务预测")
    three_fc.add_argument("--years", type=int, help="预测年数")
    three_fc.add_argument("--growth", type=float, help="收入增长率")
    three_fc.add_argument("--compact", action="store_true", help="紧凑输出")
    three_fc.set_defaults(func=cmd_three_forecast)

    # three check
    three_chk = three_sub.add_parser("check", help="配平检验")
    three_chk.add_argument("--compact", action="store_true", help="紧凑输出")
    three_chk.set_defaults(func=cmd_three_check)

    # ===== Ratio =====
    ratio_parser = subparsers.add_parser("ratio", help="财务比率分析")
    ratio_sub = ratio_parser.add_subparsers(dest="subcommand")

    # ratio calc
    ratio_calc = ratio_sub.add_parser("calc", help="计算比率")
    ratio_calc.add_argument("--type", choices=["all", "profitability", "liquidity", "solvency", "efficiency", "valuation"],
                            default="all", help="比率类型")
    ratio_calc.add_argument("--compact", action="store_true", help="紧凑输出")
    ratio_calc.set_defaults(func=cmd_ratio_calc)

    # ratio dupont
    ratio_dupont = ratio_sub.add_parser("dupont", help="杜邦分析")
    ratio_dupont.add_argument("--levels", type=int, choices=[3, 5], default=3, help="分解层级(3或5)")
    ratio_dupont.add_argument("--compact", action="store_true", help="紧凑输出")
    ratio_dupont.set_defaults(func=cmd_ratio_dupont)

    # ratio compare
    ratio_cmp = ratio_sub.add_parser("compare", help="同业对比")
    ratio_cmp.add_argument("--metrics", help="对比指标(逗号分隔)")
    ratio_cmp.add_argument("--compact", action="store_true", help="紧凑输出")
    ratio_cmp.set_defaults(func=cmd_ratio_compare)

    # ratio trend
    ratio_trend = ratio_sub.add_parser("trend", help="趋势分析")
    ratio_trend.add_argument("--metrics", help="分析指标(逗号分隔)")
    ratio_trend.add_argument("--compact", action="store_true", help="紧凑输出")
    ratio_trend.set_defaults(func=cmd_ratio_trend)

    # ===== Prepare =====
    prepare_parser = subparsers.add_parser("prepare", help="数据准备")
    prepare_sub = prepare_parser.add_subparsers(dest="subcommand")

    # prepare lbo
    prepare_lbo = prepare_sub.add_parser("lbo", help="准备LBO数据")
    prepare_lbo.add_argument("--compact", action="store_true", help="紧凑输出")
    prepare_lbo.set_defaults(func=cmd_prepare_lbo)

    # prepare ma
    prepare_ma = prepare_sub.add_parser("ma", help="准备MA数据")
    prepare_ma.add_argument("--compact", action="store_true", help="紧凑输出")
    prepare_ma.set_defaults(func=cmd_prepare_ma)

    # prepare dcf
    prepare_dcf = prepare_sub.add_parser("dcf", help="准备DCF数据")
    prepare_dcf.add_argument("--compact", action="store_true", help="紧凑输出")
    prepare_dcf.set_defaults(func=cmd_prepare_dcf)

    # ===== Tool =====
    tool_parser = subparsers.add_parser("tool", help="原子工具")
    tool_sub = tool_parser.add_subparsers(dest="subcommand")

    # tool list
    tool_list = tool_sub.add_parser("list", help="列出所有工具")
    tool_list.add_argument("--format", choices=["json", "text"], default="text", help="输出格式")
    tool_list.add_argument("--compact", action="store_true", help="紧凑输出")
    tool_list.set_defaults(func=cmd_tool_list)

    # tool run
    tool_run = tool_sub.add_parser("run", help="运行指定工具")
    tool_run.add_argument("tool_name", help="工具名称")
    tool_run.add_argument("--compact", action="store_true", help="紧凑输出")
    tool_run.set_defaults(func=cmd_tool_run)

    # tool help
    tool_help = tool_sub.add_parser("help", help="显示工具帮助")
    tool_help.add_argument("tool_name", help="工具名称")
    tool_help.set_defaults(func=cmd_tool_help)

    # 解析参数
    args = parser.parse_args()

    if args.command is None:
        parser.print_help()
        sys.exit(0)

    if hasattr(args, "func"):
        args.func(args)
    else:
        # 子命令未指定
        if args.command == "lbo":
            lbo_parser.print_help()
        elif args.command == "ma":
            ma_parser.print_help()
        elif args.command == "dcf":
            dcf_parser.print_help()
        elif args.command == "three":
            three_parser.print_help()
        elif args.command == "ratio":
            ratio_parser.print_help()
        elif args.command == "prepare":
            prepare_parser.print_help()
        elif args.command == "tool":
            tool_parser.print_help()


if __name__ == "__main__":
    main()
