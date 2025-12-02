# -*- coding: utf-8 -*-
"""
Excel 导出工具

将建模结果导出为格式化的 Excel 文件，方便人类阅读
"""

from typing import Dict, Any, List
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelWriter:
    """
    Excel 导出工具

    将建模结果导出为格式化的 Excel 文件

    使用方法:
        writer = ExcelWriter()
        writer.write_three_statement(result, "三表模型")
        writer.write_dcf(dcf_result, "DCF估值")
        writer.save("output.xlsx")
    """

    def __init__(self):
        if not HAS_OPENPYXL:
            raise ImportError("请安装 openpyxl: pip install openpyxl")

        self.wb = Workbook()
        # 删除默认的 sheet
        self.wb.remove(self.wb.active)

        # 样式定义
        self.title_font = Font(bold=True, size=14)
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        self.number_font = Font(size=11)
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _format_number(self, value, decimals=2):
        """格式化数字"""
        if isinstance(value, (int, float)):
            if abs(value) >= 10000:
                return f"{value:,.{decimals}f}"
            else:
                return f"{value:.{decimals}f}"
        return str(value)

    def _set_column_width(self, ws, col, width):
        """设置列宽"""
        ws.column_dimensions[get_column_letter(col)].width = width

    def _write_title(self, ws, row, title):
        """写标题"""
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = self.title_font
        return row + 1

    def _write_header_row(self, ws, row, headers, start_col=1):
        """写表头行"""
        for i, header in enumerate(headers):
            cell = ws.cell(row=row, column=start_col + i, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center')
        return row + 1

    def _write_data_row(self, ws, row, data, start_col=1, is_total=False):
        """写数据行"""
        for i, value in enumerate(data):
            cell = ws.cell(row=row, column=start_col + i, value=value)
            cell.border = self.thin_border
            if i == 0:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='right')
            if is_total:
                cell.font = Font(bold=True)
        return row + 1

    def write_three_statement(self, result: Dict[str, Any], sheet_name: str = "三表模型"):
        """
        写入三表模型结果

        Args:
            result: ThreeStatementModel.build() 的返回结果
            sheet_name: 工作表名称
        """
        ws = self.wb.create_sheet(title=sheet_name)

        # 设置列宽
        self._set_column_width(ws, 1, 20)
        self._set_column_width(ws, 2, 18)
        self._set_column_width(ws, 3, 40)

        row = 1

        # 元信息
        meta = result.get("_meta", {})
        if meta:
            row = self._write_title(ws, row, f"{meta.get('company', '')} 财务模型 - {meta.get('scenario', '')}")
            row += 1

        # ===== 利润表 =====
        row = self._write_title(ws, row, "利润表")
        row = self._write_header_row(ws, row, ["科目", "金额(万元)", "公式"])

        income = result.get("income_statement", {})
        income_items = [
            ("revenue", "营业收入"),
            ("cost", "营业成本"),
            ("gross_profit", "毛利"),
            ("opex", "营业费用"),
            ("ebit", "营业利润"),
            ("interest", "利息费用"),
            ("ebt", "税前利润"),
            ("tax", "所得税"),
            ("net_income", "净利润"),
        ]

        for key, label in income_items:
            if key in income:
                item = income[key]
                value = item.get("value", 0)
                formula = item.get("formula", "")
                is_total = key in ["gross_profit", "ebit", "ebt", "net_income"]
                row = self._write_data_row(ws, row, [label, value, formula], is_total=is_total)

        row += 1

        # ===== 资产负债表 =====
        row = self._write_title(ws, row, "资产负债表")
        row = self._write_header_row(ws, row, ["科目", "金额(万元)", "说明"])

        bs = result.get("balance_sheet", {})

        # 资产
        ws.cell(row=row, column=1, value="【资产】").font = Font(bold=True)
        row += 1

        assets = bs.get("assets", {})
        for key, label in [("cash", "现金"), ("receivable", "应收账款"),
                           ("inventory", "存货"), ("fixed_assets_net", "固定资产净值")]:
            if key in assets:
                item = assets[key]
                value = item.get("value", 0) if isinstance(item, dict) else item
                row = self._write_data_row(ws, row, [f"  {label}", value, ""])

        if "total_assets" in assets:
            row = self._write_data_row(ws, row, ["资产合计", assets["total_assets"].get("value", 0), ""],
                                       is_total=True)

        # 负债
        ws.cell(row=row, column=1, value="【负债】").font = Font(bold=True)
        row += 1

        liabilities = bs.get("liabilities", {})
        for key, label in [("payable", "应付账款"), ("debt", "有息负债")]:
            if key in liabilities:
                item = liabilities[key]
                value = item.get("value", 0) if isinstance(item, dict) else item
                row = self._write_data_row(ws, row, [f"  {label}", value, ""])

        if "total_liabilities" in liabilities:
            row = self._write_data_row(ws, row, ["负债合计", liabilities["total_liabilities"].get("value", 0), ""],
                                       is_total=True)

        # 权益
        ws.cell(row=row, column=1, value="【权益】").font = Font(bold=True)
        row += 1

        equity = bs.get("equity", {})
        if "total_equity" in equity:
            row = self._write_data_row(ws, row, ["权益合计", equity["total_equity"].get("value", 0), ""],
                                       is_total=True)

        row += 1

        # ===== 现金流量表 =====
        row = self._write_title(ws, row, "现金流量表")
        row = self._write_header_row(ws, row, ["科目", "金额(万元)", "公式"])

        cf = result.get("cash_flow", {})
        cf_items = [
            ("operating", "经营活动现金流"),
            ("investing", "投资活动现金流"),
            ("financing", "筹资活动现金流"),
            ("net_change", "现金净变动"),
            ("closing_cash", "期末现金"),
        ]

        for key, label in cf_items:
            if key in cf:
                item = cf[key]
                value = item.get("value", 0)
                formula = item.get("formula", "")
                is_total = key in ["net_change", "closing_cash"]
                row = self._write_data_row(ws, row, [label, value, formula], is_total=is_total)

        row += 1

        # ===== 配平检验 =====
        validation = result.get("validation", {})
        if validation:
            row = self._write_title(ws, row, "配平检验")
            status = "✓ 配平" if validation.get("is_balanced") else "✗ 不配平"
            diff = validation.get("difference", 0)
            row = self._write_data_row(ws, row, ["状态", status, f"差额: {diff:.2f}"])

    def write_dcf(self, result: Dict[str, Any], sheet_name: str = "DCF估值"):
        """
        写入 DCF 估值结果

        Args:
            result: DCFModel.valuate() 的返回结果
            sheet_name: 工作表名称
        """
        ws = self.wb.create_sheet(title=sheet_name)

        self._set_column_width(ws, 1, 20)
        self._set_column_width(ws, 2, 18)
        self._set_column_width(ws, 3, 40)

        row = 1
        row = self._write_title(ws, row, "DCF 估值模型")
        row += 1

        # 折现率
        row = self._write_title(ws, row, "折现率计算")
        row = self._write_header_row(ws, row, ["项目", "值", "公式"])

        if "cost_of_equity" in result:
            re = result["cost_of_equity"]
            row = self._write_data_row(ws, row, ["股权成本 (Re)", f"{re['value']:.2%}", re.get("formula", "")])

        if "wacc" in result:
            wacc = result["wacc"]
            row = self._write_data_row(ws, row, ["WACC", f"{wacc['value']:.2%}", wacc.get("formula", "")],
                                       is_total=True)

        row += 1

        # FCF 预测
        if "_inputs" in result and "fcf_projections" in result["_inputs"]:
            row = self._write_title(ws, row, "自由现金流预测")
            row = self._write_header_row(ws, row, ["年份", "FCF (万元)", ""])

            for year, fcf in result["_inputs"]["fcf_projections"].items():
                row = self._write_data_row(ws, row, [year, fcf, ""])

            row += 1

        # 估值结果
        row = self._write_title(ws, row, "估值结果")
        row = self._write_header_row(ws, row, ["项目", "值 (万元)", "公式"])

        items = [
            ("terminal_value", "终值"),
            ("enterprise_value", "企业价值"),
            ("equity_value", "股权价值"),
        ]

        for key, label in items:
            if key in result:
                item = result[key]
                row = self._write_data_row(ws, row, [label, item["value"], item.get("formula", "")])

        if "per_share_value" in result:
            psv = result["per_share_value"]
            row = self._write_data_row(ws, row, ["每股价值 (元)", psv["value"], psv.get("formula", "")],
                                       is_total=True)

        row += 1

        # 敏感性分析
        if "sensitivity_matrix" in result and result["sensitivity_matrix"]:
            sens = result["sensitivity_matrix"]
            row = self._write_title(ws, row, "敏感性分析 (每股价值)")

            # 获取列头
            if sens.get("data"):
                first_row = sens["data"][0]
                headers = ["WACC \\ g"] + [k for k in first_row.keys() if k != "wacc"]
                row = self._write_header_row(ws, row, headers)

                for data_row in sens["data"]:
                    values = [data_row["wacc"]] + [data_row.get(k, "N/A") for k in headers[1:]]
                    row = self._write_data_row(ws, row, values)

    def write_scenario_comparison(self, result: Dict[str, Any], sheet_name: str = "场景对比"):
        """
        写入场景对比结果

        Args:
            result: ScenarioManager.compare_scenarios() 的返回结果
            sheet_name: 工作表名称
        """
        ws = self.wb.create_sheet(title=sheet_name)

        self._set_column_width(ws, 1, 15)
        for i in range(2, 6):
            self._set_column_width(ws, i, 18)

        row = 1
        row = self._write_title(ws, row, "场景对比分析")
        row += 1

        comparison = result.get("scenario_comparison", result)

        if "headers" in comparison:
            headers = comparison["headers"]
            row = self._write_header_row(ws, row, headers)

            for data_row in comparison.get("rows", []):
                values = [data_row.get("metric", "")]
                for h in headers[1:]:
                    values.append(data_row.get(h, ""))
                row = self._write_data_row(ws, row, values)

        # 验证
        if "validation" in comparison:
            row += 1
            v = comparison["validation"]
            status = "✓ 通过" if v.get("is_valid") else "✗ 不通过"
            row = self._write_data_row(ws, row, ["场景排序验证", status, v.get("message", "")])

    def write_sensitivity_2d(self, result: Dict[str, Any], sheet_name: str = "敏感性分析"):
        """
        写入双参数敏感性矩阵

        Args:
            result: ScenarioManager.sensitivity_2d() 的返回结果
            sheet_name: 工作表名称
        """
        ws = self.wb.create_sheet(title=sheet_name)

        row = 1

        headers = result.get("headers", {})
        param1 = headers.get("rows", "参数1")
        param2 = headers.get("columns", "参数2")
        metric = result.get("output_metric", "")

        row = self._write_title(ws, row, f"敏感性分析: {metric}")
        row += 1

        if result.get("data"):
            first_row = result["data"][0]
            col_headers = [f"{param1} \\ {param2}"] + [k for k in first_row.keys() if k != param1]
            row = self._write_header_row(ws, row, col_headers)

            for data_row in result["data"]:
                p1_val = data_row.get(param1, "")
                if isinstance(p1_val, float) and p1_val < 1:
                    p1_val = f"{p1_val:.1%}"
                values = [p1_val] + [data_row.get(k, "") for k in col_headers[1:]]
                row = self._write_data_row(ws, row, values)

    def write_advanced_features(self, deferred_tax=None, impairment=None, lease=None,
                                 sheet_name: str = "高级功能"):
        """
        写入高级功能结果

        Args:
            deferred_tax: 递延税结果
            impairment: 减值测试结果
            lease: 租赁资本化结果
            sheet_name: 工作表名称
        """
        ws = self.wb.create_sheet(title=sheet_name)

        self._set_column_width(ws, 1, 25)
        self._set_column_width(ws, 2, 18)
        self._set_column_width(ws, 3, 35)

        row = 1

        # ===== 递延税 =====
        if deferred_tax:
            row = self._write_title(ws, row, "递延税计算")
            row = self._write_header_row(ws, row, ["项目", "值", "说明"])

            for scenario, data in deferred_tax.items():
                ws.cell(row=row, column=1, value=f"【{scenario}】").font = Font(bold=True)
                row += 1

                items = [
                    ("current_tax", "当期税"),
                    ("dta_change", "DTA变动"),
                    ("closing_dta", "期末DTA"),
                    ("tax_expense", "所得税费用"),
                ]
                for key, label in items:
                    if key in data:
                        item = data[key]
                        value = item.get("value", 0)
                        formula = item.get("formula", "")
                        row = self._write_data_row(ws, row, [f"  {label}", value, formula])

            row += 1

        # ===== 减值测试 =====
        if impairment:
            row = self._write_title(ws, row, "资产减值测试")
            row = self._write_header_row(ws, row, ["项目", "值", "公式"])

            test = impairment.get("impairment_test", impairment)

            items = [
                ("carrying_value", "账面价值"),
                ("recoverable_amount", "可回收金额"),
                ("impairment_loss", "减值损失"),
                ("new_carrying_value", "新账面价值"),
            ]

            for key, label in items:
                if key in test:
                    item = test[key]
                    if isinstance(item, dict):
                        value = item.get("value", 0)
                        formula = item.get("formula", "")
                    else:
                        value = item
                        formula = ""
                    is_total = key == "impairment_loss"
                    row = self._write_data_row(ws, row, [label, value, formula], is_total=is_total)

            row += 1

        # ===== 租赁资本化 =====
        if lease:
            row = self._write_title(ws, row, "租赁资本化 (IFRS 16)")

            # 初始确认
            init = lease.get("initial_recognition", {})
            if init:
                row = self._write_header_row(ws, row, ["初始确认", "值", "公式"])
                if "lease_liability" in init:
                    ll = init["lease_liability"]
                    row = self._write_data_row(ws, row, ["租赁负债", ll.get("value", 0), ll.get("formula", "")])
                if "rou_asset" in init:
                    rou = init["rou_asset"]
                    row = self._write_data_row(ws, row, ["使用权资产", rou.get("value", 0), rou.get("formula", "")])

            row += 1

            # 年度明细
            schedule = lease.get("annual_schedule", [])
            if schedule:
                row = self._write_header_row(ws, row, ["年份", "期初负债", "利息", "本金", "期末负债", "折旧"])
                for item in schedule:
                    row = self._write_data_row(ws, row, [
                        item.get("year", ""),
                        item.get("opening_liability", 0),
                        item.get("interest_expense", 0),
                        item.get("principal_payment", 0),
                        item.get("closing_liability", 0),
                        item.get("depreciation", 0),
                    ])

            row += 1

            # 影响汇总
            impact = lease.get("impact_summary", {})
            if impact:
                row = self._write_header_row(ws, row, ["影响项目", "值", ""])
                for key, label in [("old_rent_expense", "原租金费用"),
                                   ("new_depreciation", "新折旧"),
                                   ("new_interest", "新利息"),
                                   ("ebitda_improvement", "EBITDA改善"),
                                   ("net_income_impact", "净利润影响")]:
                    if key in impact:
                        row = self._write_data_row(ws, row, [label, impact[key], ""])

    def save(self, filepath: str):
        """
        保存 Excel 文件

        Args:
            filepath: 文件路径
        """
        self.wb.save(filepath)
        print(f"✓ Excel 已保存: {filepath}")
