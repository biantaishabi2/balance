#!/usr/bin/env python3
"""创建示例 Excel 文件"""

import openpyxl

wb = openpyxl.Workbook()

# 资产负债表
ws1 = wb.active
ws1.title = "资产负债表"
ws1.append(["科目", "期初", "期末"])
ws1.append(["现金", 20000, ""])
ws1.append(["应收账款", 8000, ""])
ws1.append(["存货", 12000, ""])
ws1.append(["固定资产原值", 100000, ""])
ws1.append(["累计折旧", 20000, ""])
ws1.append(["固定资产净值", "", ""])
ws1.append(["资产合计", "", ""])
ws1.append(["短期借款", 50000, ""])
ws1.append(["应付账款", 6000, ""])
ws1.append(["负债合计", "", ""])
ws1.append(["股本", 30000, ""])
ws1.append(["留存收益", 10000, ""])
ws1.append(["权益合计", "", ""])

# 损益表
ws2 = wb.create_sheet("损益表")
ws2.append(["科目", "金额"])
ws2.append(["营业收入", 100000])
ws2.append(["营业成本", 60000])
ws2.append(["毛利", ""])
ws2.append(["折旧费用", ""])
ws2.append(["其他费用", 5000])
ws2.append(["营业利润", ""])
ws2.append(["利息费用", ""])
ws2.append(["利润总额", ""])
ws2.append(["所得税", ""])
ws2.append(["净利润", ""])

# 现金流量表
ws3 = wb.create_sheet("现金流量表")
ws3.append(["科目", "金额"])
ws3.append(["经营活动现金流", ""])
ws3.append(["投资活动现金流", ""])
ws3.append(["筹资活动现金流", ""])
ws3.append(["现金净增加", ""])

# 参数
ws4 = wb.create_sheet("参数")
ws4.append(["参数", "值"])
ws4.append(["利率", 0.08])
ws4.append(["税率", 0.25])
ws4.append(["折旧年限", 10])
ws4.append(["残值", 0])
ws4.append(["分红", 5000])
ws4.append(["资本支出", 0])
ws4.append(["最低现金", 5000])

wb.save("/home/wangbo/document/balance/examples/sample.xlsx")
print("✓ 创建示例文件: examples/sample.xlsx")
