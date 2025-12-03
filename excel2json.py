#!/usr/bin/env python3
"""
excel2json - 从 Excel/CSV 提取财务数据生成标准 JSON

支持金蝶、用友、SAP 等主流财务软件导出的文件格式。

用法:
    excel2json 财务报表.xlsx --template trial_balance > tb.json
    excel2json 预算表.xlsx --template budget > budget.json
    excel2json 账龄表.xlsx --template ar_aging > ar.json

支持的模板:
    trial_balance (tb)  - 科目余额表
    budget              - 预算对比表
    ar_aging (ar)       - 应收账龄表
    ap_aging (ap)       - 应付账龄表
    voucher             - 记账凭证
    cash                - 现金流数据
"""

import sys
import json
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("错误: 需要安装 openpyxl (pip install openpyxl)", file=sys.stderr)
    sys.exit(1)

from fin_tools.tools.erp_parser import (
    parse_erp_file,
    normalize_column_name,
    COLUMN_MAPPING
)


# 旧版默认映射（保持向后兼容）
LEGACY_MAPPING = {
    "资产负债表": {
        "现金": "opening_cash",
        "应收账款": "opening_receivable",
        "存货": "opening_inventory",
        "固定资产原值": "fixed_asset_cost",
        "累计折旧": "accum_depreciation",
        "短期借款": "opening_debt",
        "应付账款": "opening_payable",
        "股本": "opening_equity",
        "留存收益": "opening_retained",
    },
    "损益表": {
        "营业收入": "revenue",
        "营业成本": "cost",
        "其他费用": "other_expense",
    },
    "参数": {
        "利率": "interest_rate",
        "税率": "tax_rate",
        "折旧年限": "fixed_asset_life",
        "残值": "fixed_asset_salvage",
        "分红": "dividend",
        "资本支出": "capex",
        "最低现金": "min_cash",
        "还款": "repayment",
        "新增股本": "new_equity",
    },
}


def find_value_by_label(sheet, label: str, value_col: int = 2) -> float | None:
    """在 sheet 中查找标签对应的值"""
    for row in sheet.iter_rows(min_row=1, max_row=100):
        cell_label = row[0].value
        if cell_label and str(cell_label).strip() == label:
            if len(row) >= value_col:
                val = row[value_col - 1].value
                if val is not None:
                    try:
                        return float(val)
                    except (ValueError, TypeError):
                        return None
    return None


def extract_legacy(filepath: str, mapping: dict = None, value_col: int = 2) -> dict:
    """旧版提取方式（保持向后兼容）"""
    if mapping is None:
        mapping = LEGACY_MAPPING

    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}

    for sheet_name, fields in mapping.items():
        target_sheet = None
        for name in wb.sheetnames:
            if sheet_name in name or name in sheet_name:
                target_sheet = wb[name]
                break

        if target_sheet is None:
            print(f"警告: 未找到 sheet '{sheet_name}'", file=sys.stderr)
            continue

        for label, json_field in fields.items():
            val = find_value_by_label(target_sheet, label, value_col)
            if val is not None:
                if "折旧" in label and val > 0:
                    val = abs(val)
                if "成本" in label or "费用" in label:
                    val = abs(val)
                result[json_field] = val
            else:
                print(f"警告: 未找到 '{sheet_name}.{label}'", file=sys.stderr)

    wb.close()
    return result


def main():
    parser = argparse.ArgumentParser(
        prog="excel2json",
        description="从 Excel/CSV 提取财务数据生成标准 JSON",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
支持的模板类型:
    trial_balance (tb)  科目余额表 - 金蝶/用友导出的科目余额表
    budget              预算对比表 - 预算 vs 实际对比
    ar_aging (ar)       应收账龄表 - 应收账款账龄分析
    ap_aging (ap)       应付账龄表 - 应付账款账龄分析
    voucher             记账凭证   - 会计凭证明细
    cash                现金流数据 - 现金流量预测输入

示例:
    # 解析金蝶导出的科目余额表
    excel2json 科目余额表.xlsx --template tb > trial_balance.json

    # 解析用友导出的预算表
    excel2json 预算表.xlsx --template budget --period 2025-11 > budget.json

    # 解析账龄表
    excel2json 应收账龄.xlsx --template ar --period 2025-12-01 > ar_aging.json

    # 旧版兼容模式（不指定 template）
    excel2json 财务报表.xlsx --col 2 > input.json
        """
    )
    parser.add_argument("excel_file", help="Excel/CSV 文件路径")
    parser.add_argument(
        "--template", "-t",
        choices=["trial_balance", "tb", "budget", "ar_aging", "ar", "ap_aging", "ap", "voucher", "cash"],
        help="模板类型（新版 ERP 解析）"
    )
    parser.add_argument(
        "--sheet", "-s",
        help="工作表名称（模糊匹配）"
    )
    parser.add_argument(
        "--header", "-H",
        type=int,
        default=1,
        help="表头行号 (默认1)"
    )
    parser.add_argument(
        "--period", "-p",
        default="",
        help="会计期间 (如 2025-11)"
    )
    parser.add_argument(
        "--company", "-C",
        default="",
        help="公司名称"
    )
    parser.add_argument(
        "--encoding", "-e",
        default="utf-8",
        help="CSV 文件编码 (默认 utf-8)"
    )
    # 旧版兼容参数
    parser.add_argument(
        "--col", "-c",
        type=int,
        default=2,
        help="数据所在列号 (旧版模式，默认2=B列)"
    )
    parser.add_argument(
        "--mapping", "-m",
        help="自定义映射配置文件 (旧版模式)"
    )
    parser.add_argument(
        "--pretty",
        action="store_true",
        default=True,
        help="格式化输出 (默认)"
    )
    parser.add_argument(
        "--compact",
        action="store_true",
        help="紧凑输出"
    )

    args = parser.parse_args()

    # 检查文件存在
    if not Path(args.excel_file).exists():
        print(f"错误: 文件不存在 '{args.excel_file}'", file=sys.stderr)
        sys.exit(1)

    # 根据是否指定 template 选择模式
    if args.template:
        # 新版 ERP 解析模式
        try:
            data = parse_erp_file(
                filepath=args.excel_file,
                template=args.template,
                sheet_name=args.sheet,
                header_row=args.header,
                period=args.period,
                company=args.company,
                encoding=args.encoding
            )
        except Exception as e:
            print(f"错误: {e}", file=sys.stderr)
            sys.exit(1)
    else:
        # 旧版兼容模式
        mapping = None
        if args.mapping:
            with open(args.mapping) as f:
                mapping = json.load(f)

        data = extract_legacy(args.excel_file, mapping, args.col)

    # 输出
    if args.compact:
        print(json.dumps(data, ensure_ascii=False))
    else:
        print(json.dumps(data, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
