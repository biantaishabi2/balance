#!/usr/bin/env python3
"""
json2excel - 将配平结果写回 Excel

用法:
    json2excel output.json 财务报表.xlsx
    cat output.json | json2excel - 财务报表.xlsx
"""

import sys
import json
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("错误: 需要安装 openpyxl (pip install openpyxl)", file=sys.stderr)
    sys.exit(1)


# 默认写入映射：JSON字段 -> Excel位置
# 格式: json_field -> (sheet名, 标签, 列号)
DEFAULT_OUTPUT_MAPPING = {
    # 资产负债表 - 期末数据（写入C列）
    "closing_cash": ("资产负债表", "现金", 3),
    "closing_receivable": ("资产负债表", "应收账款", 3),
    "closing_fixed_asset_net": ("资产负债表", "固定资产净值", 3),
    "closing_accum_depreciation": ("资产负债表", "累计折旧", 3),
    "closing_debt": ("资产负债表", "短期借款", 3),
    "closing_payable": ("资产负债表", "应付账款", 3),
    "closing_retained": ("资产负债表", "留存收益", 3),
    "closing_total_equity": ("资产负债表", "权益合计", 3),
    "total_assets": ("资产负债表", "资产合计", 3),
    "total_liabilities": ("资产负债表", "负债合计", 3),

    # 损益表
    "interest": ("损益表", "利息费用", 2),
    "depreciation": ("损益表", "折旧费用", 2),
    "gross_profit": ("损益表", "毛利", 2),
    "ebit": ("损益表", "营业利润", 2),
    "ebt": ("损益表", "利润总额", 2),
    "tax": ("损益表", "所得税", 2),
    "net_income": ("损益表", "净利润", 2),

    # 现金流量表
    "operating_cashflow": ("现金流量表", "经营活动现金流", 2),
    "investing_cashflow": ("现金流量表", "投资活动现金流", 2),
    "financing_cashflow": ("现金流量表", "筹资活动现金流", 2),
}


def find_row_by_label(sheet, label: str) -> int | None:
    """
    在 sheet 中查找标签所在行号

    Args:
        sheet: openpyxl sheet 对象
        label: 要查找的标签

    Returns:
        行号，或 None
    """
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=100), start=1):
        cell_label = row[0].value
        if cell_label and str(cell_label).strip() == label:
            return row_idx
    return None


def write_to_excel(data: dict, filepath: str, mapping: dict = None) -> dict:
    """
    将数据写入 Excel

    Args:
        data: 要写入的数据字典
        filepath: Excel 文件路径
        mapping: 写入映射配置

    Returns:
        写入结果统计
    """
    if mapping is None:
        mapping = DEFAULT_OUTPUT_MAPPING

    # 检查文件是否存在
    if Path(filepath).exists():
        wb = openpyxl.load_workbook(filepath)
    else:
        # 创建新文件
        wb = openpyxl.Workbook()
        # 创建默认 sheet
        for sheet_name in ["资产负债表", "损益表", "现金流量表"]:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
        # 删除默认的 Sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    stats = {"written": 0, "skipped": 0, "not_found": 0}

    for json_field, (sheet_name, label, col) in mapping.items():
        # 检查字段是否在数据中
        if json_field not in data:
            stats["skipped"] += 1
            continue

        value = data[json_field]

        # 查找目标 sheet
        target_sheet = None
        for name in wb.sheetnames:
            if sheet_name in name or name in sheet_name:
                target_sheet = wb[name]
                break

        if target_sheet is None:
            # 创建 sheet
            target_sheet = wb.create_sheet(sheet_name)
            print(f"创建 sheet: {sheet_name}", file=sys.stderr)

        # 查找行号
        row = find_row_by_label(target_sheet, label)

        if row is None:
            # 找不到标签，追加到末尾
            row = target_sheet.max_row + 1
            target_sheet.cell(row=row, column=1, value=label)
            print(f"新增行: {sheet_name}.{label}", file=sys.stderr)

        # 写入值
        target_sheet.cell(row=row, column=col, value=value)
        stats["written"] += 1

    # 保存
    wb.save(filepath)
    wb.close()

    return stats


def main():
    parser = argparse.ArgumentParser(
        description="将配平结果写回 Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
    json2excel output.json 财务报表.xlsx
    cat output.json | json2excel - 财务报表.xlsx
    json2excel output.json 财务报表.xlsx --mapping my_mapping.json

说明:
    - 如果 Excel 文件不存在，会创建新文件
    - 如果找不到对应的行，会追加到末尾
    - 默认将期末数据写入 C 列，损益数据写入 B 列
        """
    )
    parser.add_argument(
        "json_file",
        help="JSON 文件路径，或 '-' 表示从 stdin 读取"
    )
    parser.add_argument("excel_file", help="Excel 文件路径")
    parser.add_argument(
        "--mapping", "-m",
        help="自定义映射配置文件 (JSON格式)"
    )
    parser.add_argument(
        "--quiet", "-q",
        action="store_true",
        help="静默模式，不输出统计信息"
    )

    args = parser.parse_args()

    # 读取 JSON
    if args.json_file == "-":
        data = json.load(sys.stdin)
    else:
        if not Path(args.json_file).exists():
            print(f"错误: 文件不存在 '{args.json_file}'", file=sys.stderr)
            sys.exit(1)
        with open(args.json_file) as f:
            data = json.load(f)

    # 加载自定义映射
    mapping = None
    if args.mapping:
        with open(args.mapping) as f:
            mapping = json.load(f)

    # 写入 Excel
    stats = write_to_excel(data, args.excel_file, mapping)

    if not args.quiet:
        print(f"完成: 写入 {stats['written']} 个字段, 跳过 {stats['skipped']} 个", file=sys.stderr)


if __name__ == "__main__":
    main()
