#!/usr/bin/env python3
"""
excel_inspect - 查看 Excel 结构，帮助生成映射配置

用法:
    excel_inspect 报表.xlsx              # 查看所有 sheet 结构
    excel_inspect 报表.xlsx --sheet 资产  # 只看某个 sheet
    excel_inspect 报表.xlsx --json       # 输出 JSON 格式（给 AI 用）
"""

import sys
import json
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("错误: 需要安装 openpyxl", file=sys.stderr)
    sys.exit(1)


def inspect_excel(filepath: str, target_sheet: str = None, max_rows: int = 30) -> dict:
    """
    检查 Excel 结构

    Returns:
        {
            "sheets": ["sheet1", "sheet2"],
            "structure": {
                "sheet1": {
                    "headers": ["A列标题", "B列标题"],
                    "rows": [
                        {"row": 2, "A": "现金", "B": 10000},
                        ...
                    ]
                }
            }
        }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {
        "file": filepath,
        "sheets": wb.sheetnames,
        "structure": {}
    }

    for sheet_name in wb.sheetnames:
        if target_sheet and target_sheet not in sheet_name:
            continue

        sheet = wb[sheet_name]
        sheet_data = {
            "total_rows": sheet.max_row,
            "total_cols": sheet.max_column,
            "headers": [],
            "label_column": [],  # 第一列的所有标签
            "sample_data": []
        }

        # 读取表头（第一行）
        for col in range(1, min(sheet.max_column + 1, 10)):
            val = sheet.cell(row=1, column=col).value
            sheet_data["headers"].append(str(val) if val else f"列{col}")

        # 读取第一列（通常是科目名/标签）
        for row in range(1, min(sheet.max_row + 1, max_rows + 1)):
            val = sheet.cell(row=row, column=1).value
            if val:
                sheet_data["label_column"].append({
                    "row": row,
                    "label": str(val).strip()
                })

        # 读取样本数据（前N行）
        for row in range(2, min(sheet.max_row + 1, max_rows + 1)):
            row_data = {"row": row}
            has_data = False
            for col in range(1, min(sheet.max_column + 1, 10)):
                val = sheet.cell(row=row, column=col).value
                col_letter = openpyxl.utils.get_column_letter(col)
                if val is not None:
                    row_data[col_letter] = val
                    has_data = True
            if has_data:
                sheet_data["sample_data"].append(row_data)

        result["structure"][sheet_name] = sheet_data

    wb.close()
    return result


def print_human_readable(data: dict):
    """人类可读格式输出"""
    print(f"文件: {data['file']}")
    print(f"Sheets: {', '.join(data['sheets'])}")
    print()

    for sheet_name, sheet_data in data["structure"].items():
        print(f"═══ {sheet_name} ═══")
        print(f"  大小: {sheet_data['total_rows']} 行 × {sheet_data['total_cols']} 列")
        print(f"  表头: {sheet_data['headers']}")
        print()
        print("  第一列标签（科目名）:")
        for item in sheet_data["label_column"][:20]:
            print(f"    行{item['row']}: {item['label']}")
        if len(sheet_data["label_column"]) > 20:
            print(f"    ... 还有 {len(sheet_data['label_column']) - 20} 行")
        print()


def main():
    parser = argparse.ArgumentParser(
        description="查看 Excel 结构，帮助生成映射配置",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
    excel_inspect 报表.xlsx
    excel_inspect 报表.xlsx --sheet 资产
    excel_inspect 报表.xlsx --json

AI 使用流程:
    1. excel_inspect 用户文件.xlsx --json  # 查看结构
    2. 根据结构生成 mapping.json
    3. excel2json 用户文件.xlsx --mapping mapping.json | balance
        """
    )
    parser.add_argument("excel_file", help="Excel 文件路径")
    parser.add_argument(
        "--sheet", "-s",
        help="只查看包含此关键词的 sheet"
    )
    parser.add_argument(
        "--json", "-j",
        action="store_true",
        help="输出 JSON 格式"
    )
    parser.add_argument(
        "--rows", "-r",
        type=int,
        default=30,
        help="最多读取行数 (默认30)"
    )

    args = parser.parse_args()

    if not Path(args.excel_file).exists():
        print(f"错误: 文件不存在 '{args.excel_file}'", file=sys.stderr)
        sys.exit(1)

    data = inspect_excel(args.excel_file, args.sheet, args.rows)

    if args.json:
        print(json.dumps(data, indent=2, ensure_ascii=False, default=str))
    else:
        print_human_readable(data)


if __name__ == "__main__":
    main()
